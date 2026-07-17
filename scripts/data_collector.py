"""
Data Collector Module v2.0
==========================

负责从 7 大数据源采集数据:
1. Amazon 商品数据 (amazon-product-scraper 优先，Apify 补缺)
2. Amazon 差评数据 (amazon-product-scraper 优先，Apify 补缺)
3. TikTok 流量数据 (Apify)
4. Reddit 用户讨论 (Apify Reddit Scraper)    [NEW]
5. Google Trends 关键词趋势                    [NEW]
6. 1688 供应商成本 (Apify)
7. 市场报告数据 (Apify Web Scraper)

Amazon 优先使用低成本本地 scraper；其他平台优先通过 Apify API 完成。
采集失败的数据保持为空，不生成 mock 数据。
"""

import os
import json
import time
import re


def _run_attr(run, key, default=None):
    """兼容 apify_client v3.x Run 对象和 dict 格式"""
    # v3.x: Run 对象用属性 (run.id, run.status, run.default_dataset_id)
    attr_map = {
        "id": "id",
        "status": "status",
        "defaultDatasetId": "default_dataset_id",
        "errorMessage": "error_message",
    }
    attr_name = attr_map.get(key, key)
    if hasattr(run, attr_name):
        val = getattr(run, attr_name, default)
        return val if val is not None else default
    # fallback: dict 格式
    if isinstance(run, dict):
        return run.get(key, default)
    return default
import subprocess
import sys
from typing import List, Dict, Optional, Any
from pathlib import Path
from collections import Counter

from config import CONFIG, MARKET_MAP


class DataCollector:
    """统一数据采集器"""

    def __init__(self, market: str, config: dict):
        self.market = market.upper()
        self.config = config
        self.collection_config = config.get("collection", {})
        self.api_token = os.getenv("APIFY_API_TOKEN", config.get("apify_token", ""))
        self.market_config = MARKET_MAP.get(self.market, MARKET_MAP["US"])
        self._apify_client = None

    @property
    def apify_client(self):
        """懒加载 Apify 客户端"""
        if self._apify_client is None and self.api_token:
            try:
                from apify_client import ApifyClient
                self._apify_client = ApifyClient(self.api_token)
            except ImportError:
                print("⚠️ 请安装 apify-client: pip install apify-client")
            except Exception as e:
                print(f"⚠️ Apify 客户端初始化失败: {e}")
        return self._apify_client

    # ================================================================
    # 1. Amazon 商品数据采集
    # ================================================================

    def fetch_amazon_products(self, keywords: List[str]) -> List[Dict]:
        """采集 Amazon 商品数据"""
        products = []
        for keyword in keywords[:1]:
            print(f"      采集关键词: {keyword}")
            result = self._collect_amazon_local_keyword(keyword)
            min_items = max(50, int(self.collection_config.get("min_amazon_products", 50)))
            if len(result) < min_items:
                supplement = self._collect_amazon_keyword(keyword)
                by_asin = {str(item.get("asin")): item for item in result if item.get("asin")}
                for item in supplement:
                    asin = str(item.get("asin", ""))
                    if asin and asin not in by_asin:
                        by_asin[asin] = item
                result = list(by_asin.values())
            products.extend(result)
            time.sleep(self.collection_config.get("request_delay", 2))
        return products

    def _collect_amazon_local_keyword(self, keyword: str) -> List[Dict]:
        """优先通过 sibling amazon-product-scraper 低成本采集 Amazon 数据。"""
        scraper_script = Path(__file__).resolve().parents[2] / "amazon-product-scraper" / "scripts" / "amazon_scraper_core.py"
        if not scraper_script.exists():
            return []

        output_dir = Path(self.collection_config.get("output_dir", Path(__file__).resolve().parents[1] / "output" / "amazon_local"))
        output_dir.mkdir(parents=True, exist_ok=True)
        requested = int(self.collection_config.get("max_products", 100))
        max_items = min(max(requested, 50), 100)

        before = set(output_dir.glob("amazon_*.xlsx"))
        cmd = [
            sys.executable,
            str(scraper_script),
            keyword,
            "sales",
            str(max_items),
            str(output_dir),
            "reviews",
            "2",
        ]

        try:
            subprocess.run(cmd, check=True, capture_output=True, text=True, timeout=900)
        except Exception as exc:
            print(f"      ⚠️ 本地 Amazon scraper 未成功，改用 Apify 补缺: {str(exc)[:120]}")
            return []

        after = set(output_dir.glob("amazon_*.xlsx"))
        new_files = sorted(after - before, key=lambda p: p.stat().st_mtime, reverse=True)
        if not new_files:
            new_files = sorted(output_dir.glob("amazon_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        if not new_files:
            return []

        products = self._read_local_amazon_excel(new_files[0], keyword)
        if products:
            print(f"      ✅ Amazon 本地 scraper: {len(products)} 个商品 (关键词: {keyword})")
        return products

    def _read_local_amazon_excel(self, excel_path: Path, keyword: str) -> List[Dict]:
        """读取 amazon-product-scraper 导出的 Excel。"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("      ⚠️ openpyxl 未安装，无法读取本地 Amazon scraper 输出")
            return []

        try:
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active
            headers = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], 1) if cell.value}

            def cell(row, name, default=""):
                idx = headers.get(name)
                return row[idx - 1].value if idx and idx <= len(row) else default

            products = []
            for row in ws.iter_rows(min_row=2):
                asin = cell(row, "ASIN")
                title = cell(row, "商品标题")
                if not asin or str(asin).lower() == "null" or not title:
                    continue
                products.append({
                    "asin": str(asin),
                    "title": str(title),
                    "brand": cell(row, "品牌名称", "") or "",
                    "price": self._to_float(cell(row, "价格($)", 0)),
                    "rating": self._to_float(cell(row, "评分", 0)),
                    "review_count": self._to_int(cell(row, "评论数", 0)),
                    "bsr": cell(row, "BSR排名", "") or "",
                    "bullets": cell(row, "五点描述", "") or "",
                    "specifications": cell(row, "规格参数", "") or "",
                    "estimated_monthly_sales": cell(row, "月销量", "") or "",
                    "negative_review_summary": cell(row, "客户差评", "") or "",
                    "image": cell(row, "首图", "") or "",
                    "keyword": keyword,
                    "market": self.market,
                    "source": "amazon-product-scraper",
                    "source_file": str(excel_path),
                    "source_url": f"https://www.{self.market_config.get('domain', 'amazon.com')}/dp/{str(asin)}",
                })
            return products
        except Exception as exc:
            print(f"      ⚠️ 本地 Amazon Excel 读取失败: {str(exc)[:120]}")
            return []

    @staticmethod
    def _to_float(value: Any) -> float:
        if value in (None, ""):
            return 0.0
        match = re.search(r"\d+(?:\.\d+)?", str(value).replace(",", ""))
        return float(match.group(0)) if match else 0.0

    @staticmethod
    def _to_int(value: Any) -> int:
        if value in (None, ""):
            return 0
        match = re.search(r"\d+", str(value).replace(",", ""))
        return int(match.group(0)) if match else 0

    def _collect_amazon_keyword(self, keyword: str) -> List[Dict]:
        """通过 Apify 采集单个 Amazon 关键词
        
        v3.1: 修复 Actor 参数 — 使用 categoryUrls (非 searchUrls)
        Actor: junglee/free-amazon-product-scraper ✅ 已验证
        """
        client = self.apify_client
        if not client:
            print(f"      ⚠️ Apify 客户端未初始化，跳过 Amazon 采集")
            return []

        try:
            domain = self.market_config.get("domain", "amazon.com")
            search_url = f"https://www.{domain}/s?k={keyword.replace(' ', '+')}&s=review-rank"
            requested = int(self.collection_config.get("max_products", 100))
            max_items = min(max(requested, 50), 100)

            run_input = {
                "categoryUrls": [{"url": search_url}],
                "maxItems": max_items,
                "proxyConfiguration": {"useApifyProxy": True},
            }

            actor_id = "junglee/free-amazon-product-scraper"
            
            # 启动 run (先用 wait_for_finish，再用轮询兜底)
            try:
                run = client.actor(actor_id).start(
                    run_input=run_input,
                    wait_for_finish=300
                )
            except TypeError:
                actor_call = client.actor(actor_id).start(run_input=run_input)
                run = client.actor(actor_call["id"]).wait_for_finish(timeout_secs=300)
            
            # v3.1 fix: apify_client v3.x 返回 Run 对象而非 dict，兼容两种格式
            run_id = _run_attr(run, "id")
            run_status = _run_attr(run, "status", "UNKNOWN")
            default_dataset_id = _run_attr(run, "defaultDatasetId")
            
            # 如果 wait_for_finish 返回了非终态（READY/RUNNING），轮询等待
            poll_attempts = 0
            while run_status in ("READY", "RUNNING") and poll_attempts < 40:
                time.sleep(15)
                run_obj = client.run(run_id).get()
                run_status = _run_attr(run_obj, "status", "UNKNOWN")
                poll_attempts += 1
            
            if run_status != "SUCCEEDED":
                error_msg = _run_attr(run, "errorMessage", "No error detail") or "No error detail"
                print(f"      ⚠️ Amazon 采集未成功 (status={run_status}): {str(error_msg)[:120]}")
                return []
            
            dataset_items = list(client.dataset(default_dataset_id).iterate_items())
            print(f"      ✅ Amazon: {len(dataset_items)} 个商品 (关键词: {keyword})")
            return [self._transform_amazon_item(item, keyword) for item in dataset_items]

        except Exception as e:
            print(f"      ⚠️ Apify Amazon 采集失败: {str(e)[:120]}")
            return []

    def _transform_amazon_item(self, item: dict, keyword: str) -> dict:
        """转换 Apify 返回的 Amazon 数据"""
        main_image = ""
        if item.get("images"):
            imgs = item["images"]
            if isinstance(imgs, list) and len(imgs) > 0:
                main_image = imgs[0].get("link", "") if isinstance(imgs[0], dict) else str(imgs[0])

        price = ""
        price_data = item.get("price", {})
        if isinstance(price_data, dict):
            price = price_data.get("value", "")
        elif price_data:
            price = str(price_data)

        bsr_text = ""
        bsr_list = item.get("bestsellerRanks", [])
        if isinstance(bsr_list, list):
            bsr_text = " | ".join([f"#{r.get('rank','')} in {r.get('category','')}" for r in bsr_list[:3]])

        domain = self.market_config.get("domain", "amazon.com")
        return {
            "asin": item.get("asin", ""),
            "title": item.get("title", ""),
            "brand": item.get("brand", ""),
            "price": float(price) if price else 0,
            "rating": float(item.get("stars", 0)),
            "review_count": int(item.get("reviewsCount", 0)),
            "bsr": bsr_text,
            "image": main_image,
            "category": item.get("breadCrumbs", item.get("category", "")),
            "bullets": item.get("features", item.get("bulletPoints", [])),
            "launch_date": item.get("dateFirstAvailable", ""),
            "variation_count": len(item.get("variantAsins", []) or []),
            "estimated_monthly_sales": item.get("estimatedMonthlySales", item.get("sales", "")),
            "keyword": keyword,
            "market": self.market,
            "source": "apify-amazon-scraper",
            # 来源链接：每条商品对应 Amazon 详情页，便于人工核对
            "source_url": f"https://www.{domain}/dp/{item.get('asin', '')}" if item.get("asin") else "",
        }

    # ================================================================
    # 2. Amazon 差评采集 (NEW)
    # ================================================================

    def fetch_negative_reviews(self, asins: List[str], max_per_product: int = 50) -> Dict:
        """
        采集竞品差评（1-3星评论）

        Args:
            asins: 商品 ASIN 列表 (Top 20 竞品)
            max_per_product: 每个商品最多采集多少条差评

        Returns:
            差评分析结果字典
        """
        print("   ├─ 采集竞品差评数据...")
        client = self.apify_client

        if not client:
            print("      ⚠️ Apify 客户端未初始化，跳过差评采集")
            return {"total_reviews": 0, "pain_points": [], "pain_categories": {}, "top_negative_keywords": [], "raw_reviews": []}
        
        if not asins:
            print("      ⚠️ 无 ASIN 可采集，跳过差评")
            return {"total_reviews": 0, "pain_points": [], "pain_categories": {}, "top_negative_keywords": [], "raw_reviews": []}

        actor_id = "junglee/amazon-reviews-scraper"

        try:
            domain = self.market_config.get("domain", "amazon.com")
            product_urls = [{"url": f"https://www.{domain}/dp/{asin}"} for asin in asins[:20]]

            run_input = {
                "productUrls": product_urls,
                "maxReviews": max_per_product,
                "filterByRatings": ["oneStar", "twoStar", "threeStar"],
                "scrapeReviewerInfo": False,
                "includeAnswers": False,
            }

            try:
                run = client.actor(actor_id).start(
                    run_input=run_input,
                    wait_for_finish=180
                )
            except TypeError:
                actor_call = client.actor(actor_id).start(run_input=run_input)
                run = client.actor(actor_call["id"]).wait_for_finish(timeout_secs=180)

            # v3.1 fix: 轮询兜底
            run_id = _run_attr(run, "id")
            poll_attempts = 0
            while _run_attr(run, "status", "UNKNOWN") in ("READY", "RUNNING") and poll_attempts < 20:
                time.sleep(10)
                run_obj = client.run(run_id).get()
                poll_attempts += 1

            if _run_attr(run, "status", "UNKNOWN") != "SUCCEEDED":
                print(f"      ⚠️ 差评采集未成功 (status={_run_attr(run, 'status')})")
                return self._analyze_reviews([])

            dataset_items = list(client.dataset(_run_attr(run, "defaultDatasetId")).iterate_items())
            all_reviews = []
            for item in dataset_items:
                review_url = item.get("productUrl") or item.get("url") or ""
                asin_match = re.search(r"/dp/([A-Z0-9]{10})", review_url)
                review_asin = asin_match.group(1) if asin_match else item.get("asin", "")
                domain = self.market_config.get("domain", "amazon.com")
                all_reviews.append({
                    "asin": review_asin,
                    "title": item.get("reviewTitle", ""),
                    "text": item.get("reviewDescription", ""),
                    "rating": item.get("ratingScore", "3"),
                    "date": item.get("date", ""),
                    "verified": item.get("isVerified", False),
                    # 来源链接：每条差评对应其商品详情页，便于人工核对
                    "source_url": f"https://www.{domain}/dp/{review_asin}" if review_asin else "",
                })
        except Exception as e:
            print(f"      ⚠️ 差评采集失败: {str(e)[:100]}")
            return self._analyze_reviews([])

        print(f"      ✅ 差评: {len(all_reviews)} 条 ({len(asins[:20])} 个 ASIN)")
        return self._analyze_reviews(all_reviews)

    def _analyze_reviews(self, reviews: List[Dict]) -> Dict:
        """对采集到的差评进行 NLP 分析"""
        if not reviews:
            return {"total_reviews": 0, "pain_points": [], "pain_categories": {}, "top_negative_keywords": [], "raw_reviews": []}

        # 合并所有评论文本
        all_text = " ".join([r.get("title", "") + " " + r.get("text", "") for r in reviews])

        # 提取高频词 (简化版，实际应使用 jieba 或 nltk)
        words = self._extract_keywords(all_text)

        # 痛点分类
        categories = self._categorize_pain_points(words, all_text)

        # 提取 Top 负面关键词
        pain_points = [
            {"word": w, "weight": c}
            for w, c in words.most_common(80)
        ]

        return {
            "total_reviews": len(reviews),
                "pain_points": pain_points,
                "pain_categories": categories,
                "top_negative_keywords": [w for w, _ in words.most_common(20)],
                "raw_sample": [r.get("text", "")[:200] for r in reviews[:5]],
                "raw_reviews": reviews,
            }

    def _extract_keywords(self, text: str) -> Counter:
        """提取关键词及频率"""
        # 负面关键词词典
        negative_patterns = [
            r'\b(broken?|damaged?|defective|cheap|flimsy|poor quality)\b',
            r'\b(doesn\'?t work|stopped working|useless|waste of money)\b',
            r'\b(too small|too big|doesn\'?t fit|wrong size)\b',
            r'\b(not as described|misleading|different from|not what I expected)\b',
            r'\b(leaking|spilled|melted|sticky|dirty)\b',
            r'\b(bad smell|smells|stinks|chemical)\b',
            r'\b(hard to use|difficult|complicated|confusing)\b',
            r'\b(returned|refund|disappointed|regret)\b',
            r'\b(dangerous|unsafe|hurt|injured)\b',
            r'\b(missing|incomplete|missing parts)\b',
            r'\b(overpriced|not worth|expensive for)\b',
            r'\b(fake|counterfeit|knockoff)\b',
        ]

        word_counts = Counter()
        text_lower = text.lower()

        for pattern in negative_patterns:
            matches = re.findall(pattern, text_lower, re.IGNORECASE)
            for match in matches:
                if isinstance(match, tuple):
                    match = match[0]
                word_counts[match] += 1

        # 如果没有匹配到，使用简单分词
        if not word_counts:
            simple_words = re.findall(r'\b[a-z]{{3,}}\b', text_lower)
            word_counts = Counter(simple_words)

        return word_counts

    def _categorize_pain_points(self, words: Counter, text: str) -> Dict[str, int]:
        """将痛点分类"""
        categories = {
            "产品质量": 0,
            "功能问题": 0,
            "包装问题": 0,
            "服务问题": 0,
            "期望落差": 0,
        }

        quality_words = {"broken", "damaged", "defective", "cheap", "flimsy", "poor quality", "fake", "counterfeit"}
        function_words = {"doesn't work", "stopped working", "useless", "waste of money", "hard to use", "difficult"}
        package_words = {"leaking", "spilled", "melted", "dirty", "missing", "incomplete"}
        service_words = {"returned", "refund", "late", "customer service"}
        expectation_words = {"not as described", "misleading", "different from", "not what i expected", "overpriced"}

        for word, count in words.items():
            w = word.lower()
            if w in quality_words:
                categories["产品质量"] += count
            elif w in function_words:
                categories["功能问题"] += count
            elif w in package_words:
                categories["包装问题"] += count
            elif w in service_words:
                categories["服务问题"] += count
            elif w in expectation_words:
                categories["期望落差"] += count

        return categories

    # ================================================================
    # 3. TikTok 数据采集
    # ================================================================

    def fetch_tiktok_data(self, keywords: List[str]) -> List[Dict]:
        """采集 TikTok 视频级原始数据，去重后目标不少于 50 条。"""
        print("   ├─ 采集 TikTok 流量数据...")
        client = self.apify_client

        if not client:
            print("      ⚠️ Apify 客户端未初始化，跳过 TikTok")
            return []

        queries = keywords[:1]
        tiktok_data = []
        actor_id = os.getenv("APIFY_TIKTOK_ACTOR_ID", "clockworks/tiktok-scraper")

        for query in queries:
            try:
                min_items = int(self.config.get("tiktok", {}).get("min_videos", 50))
                requested = int(self.config.get("tiktok", {}).get("max_videos", 50))
                run_input = {
                    "searchQueries": [query],
                    "searchSection": "",
                    "resultsPerPage": max(min_items, requested, 50),
                    "downloadSubtitlesOptions": "NEVER_DOWNLOAD_SUBTITLES",
                    "proxyConfiguration": {
                        "useApifyProxy": True
                    }
                }
                
                try:
                    run = client.actor(actor_id).start(
                        run_input=run_input,
                        wait_for_finish=180
                    )
                except TypeError:
                    actor_call = client.actor(actor_id).start(run_input=run_input)
                    run = client.actor(actor_call["id"]).wait_for_finish(timeout_secs=180)
                
                # v3.1 fix: 轮询兜底
                run_id = _run_attr(run, "id")
                poll_attempts = 0
                while _run_attr(run, "status", "UNKNOWN") in ("READY", "RUNNING") and poll_attempts < 20:
                    time.sleep(10)
                    run_obj = client.run(run_id).get()
                    poll_attempts += 1
                
                if _run_attr(run, "status", "UNKNOWN") != "SUCCEEDED":
                    print(f"      ⚠️ TikTok 搜索 {query} 采集未成功 (status={_run_attr(run, 'status')})")
                    continue
                
                dataset_items = list(client.dataset(_run_attr(run, "defaultDatasetId")).iterate_items())
                seen = set()
                for item in dataset_items:
                    meta = item.get("videoMeta") or {}
                    author = item.get("authorMeta") or {}
                    video_url = item.get("webVideoUrl") or item.get("url") or ""
                    video_id = item.get("id") or item.get("videoId") or video_url
                    if not video_id or video_id in seen:
                        continue
                    seen.add(video_id)
                    tiktok_data.append({
                        "video_id": video_id,
                        "description": item.get("text", ""),
                        "author": author.get("name", ""),
                        "author_url": author.get("profileUrl", ""),
                        "duration": meta.get("duration", ""),
                        "views": item.get("playCount", 0),
                        "likes": item.get("diggCount", 0),
                        "comments": item.get("commentCount", 0),
                        "shares": item.get("shareCount", 0),
                        "saves": item.get("collectCount", 0),
                        "hashtags": [h.get("name", "") for h in (item.get("hashtags") or [])],
                        "create_time": item.get("createTimeISO", item.get("createTime", "")),
                        "music": (item.get("musicMeta") or {}).get("musicName", ""),
                        "keyword": query,
                        "source": actor_id,
                        "run_id": _run_attr(run, "id"),
                        "dataset_id": _run_attr(run, "defaultDatasetId"),
                        "source_url": video_url,
                    })
                print(f"      ✅ TikTok {query}: {len(tiktok_data)} 条去重视频")

                time.sleep(3)
            except Exception as e:
                print(f"      ⚠️ TikTok 采集失败 ({query}): {str(e)[:120]}")

        return tiktok_data

    # ================================================================
    # 3b. TikTok Shop 销量数据采集 (NEW - 用户要求)
    # ================================================================

    def fetch_tiktok_shop_sales(self, keywords: List[str]) -> List[Dict]:
        """采集 TikTok Shop 该品类销量数据（GMV/销量/价格/店铺）

        Actor: 优先复用 TikTok Shop Scraper（如 clockworks/tiktok-shop-scraper 或
        galaxy.hk/tiktok-shop-scraper 等已验证 Actor）；不可用则返回空列表并记录状态，
        由 Web Search 兜底补充公开销量估算。
        """
        print("   ├─ 采集 TikTok Shop 销量数据...")
        client = self.apify_client
        if not client:
            print("      ⚠️ Apify 客户端未初始化，跳过 TikTok Shop 销量")
            return []

        # 候选 Actor（按优先级尝试）
        shop_actors = [
            os.getenv("APIFY_TIKTOK_SHOP_ACTOR_ID", "").strip(),
            "clockworks/tiktok-shop-scraper",
            "galaxy.hk/tiktok-shop-scraper",
            "microworld/tiktok-shop-scraper",
        ]
        shop_actors = list(dict.fromkeys(actor for actor in shop_actors if actor))

        shop_data = []
        for actor_id in shop_actors:
            for kw in keywords[:1]:
                try:
                    run_input = {
                        "searchKeywords": [kw],
                        "maxItems": 30,
                        "proxyConfiguration": {"useApifyProxy": True},
                    }
                    try:
                        run = client.actor(actor_id).start(run_input=run_input, wait_for_finish=180)
                    except TypeError:
                        actor_call = client.actor(actor_id).start(run_input=run_input)
                        run = client.actor(actor_call["id"]).wait_for_finish(timeout_secs=180)

                    if _run_attr(run, "status", "UNKNOWN") != "SUCCEEDED":
                        continue

                    dataset_items = list(client.dataset(_run_attr(run, "defaultDatasetId")).iterate_items())
                    for item in dataset_items:
                        product_url = item.get("productUrl") or item.get("url") or ""
                        shop_data.append({
                            "keyword": kw,
                            "product_title": item.get("title", ""),
                            "price": item.get("price", 0),
                            "sales_volume": item.get("sales", item.get("salesVolume", 0)),
                            "gmv": item.get("gmv", 0),
                            "shop_name": item.get("shopName", item.get("sellerName", "")),
                            "rating": item.get("rating", 0),
                            "source_url": product_url,
                        })
                    time.sleep(2)
                except Exception:
                    continue
            if shop_data:
                break

        if not shop_data:
            print("      ⚠️ TikTok Shop 无可用 Actor，销量数据由 Web Search 兜底")
            return []

        print(f"      ✅ TikTok Shop: {len(shop_data)} 个商品销量数据")
        return shop_data

    # ================================================================
    # 4. Reddit 用户讨论采集 (NEW)
    # ================================================================

    def fetch_reddit_data(self, keywords: List[str], subreddits: List[str] = None) -> Dict:
        """
        采集 Reddit 用户讨论数据
        
        v3.1: trudax/reddit-scraper 已转为付费，免费 Actor 不可用。
        改用空返回 + 标记状态，由 Web Search 层补充。
        """
        print("   ├─ 采集 Reddit 用户讨论...")
        client = self.apify_client
        actor_id = os.getenv("APIFY_REDDIT_ACTOR_ID", "trudax/reddit-scraper")

        if not client:
            print("      ⚠️ Apify 客户端未初始化，Reddit 数据由 Web Search 补充")
            return {"total_posts": 0, "sentiment": {}, "top_topics": {}, "demand_gaps": [], "raw_posts": [], "_source": "unavailable"}

        if subreddits is None:
            subreddits = self._get_relevant_subreddits()

        all_posts = []

        for kw in keywords[:1]:
            try:
                run_input = {
                    "searchTerms": [kw],
                    "maxPosts": 50,
                    "subreddits": subreddits,
                    "proxyConfiguration": {"useApifyProxy": True}
                }

                try:
                    run = client.actor(actor_id).start(
                        run_input=run_input,
                        wait_for_finish=120
                    )
                except TypeError:
                    actor_call = client.actor(actor_id).start(run_input=run_input)
                    run = client.actor(actor_call["id"]).wait_for_finish(timeout_secs=120)
                
                if _run_attr(run, "status", "UNKNOWN") != "SUCCEEDED":
                    continue
                
                dataset_items = list(client.dataset(_run_attr(run, "defaultDatasetId")).iterate_items())

                for item in dataset_items:
                    all_posts.append({
                        "title": item.get("title", ""),
                        "text": item.get("body", "")[:500],
                        "subreddit": item.get("subreddit", ""),
                        "score": item.get("score", 0),
                        "num_comments": item.get("numberOfComments", 0),
                        "url": item.get("url", ""),
                    })

                time.sleep(3)
            except Exception as e:
                print(f"      ⚠️ Reddit 采集失败: {str(e)[:120]}")
                break

        if not all_posts:
            print("      ⚠️ Reddit 无数据 (Actor 不可用或已转为付费)")
            return {"total_posts": 0, "sentiment": {}, "top_topics": {}, "demand_gaps": [], "raw_posts": [], "_source": "unavailable"}

        print(f"      ✅ Reddit: {len(all_posts)} 篇帖子")
        return self._analyze_reddit_posts(all_posts)

    def _get_relevant_subreddits(self) -> List[str]:
        """根据品类获取相关子版块"""
        category = self.metadata.get("category", "").lower() if hasattr(self, 'metadata') else ""
        default_subs = [
            "AmazonReviews", "FulfillmentByAmazon", "ecommerce",
            "ProductReviews", "BuyItForLife", "shoppingaddiction"
        ]
        return default_subs

    def _analyze_reddit_posts(self, posts: List[Dict]) -> Dict:
        """分析 Reddit 帖子"""
        if not posts:
            return {"total_posts": 0, "sentiment": {}, "top_topics": {}, "demand_gaps": []}

        # 情绪分析 (基于简单关键词)
        positive_words = ["love", "great", "best", "recommend", "excellent", "perfect", "amazing"]
        negative_words = ["hate", "worst", "terrible", "avoid", "waste", "disappointed", "regret", "broke"]

        sentiment = {"正面": 0, "中性": 0, "负面": 0}
        topic_counts = Counter()

        for post in posts:
            text = (post.get("title", "") + " " + post.get("text", "")).lower()

            pos = sum(1 for w in positive_words if w in text)
            neg = sum(1 for w in negative_words if w in text)

            if neg > pos:
                sentiment["负面"] += 1
            elif pos > neg:
                sentiment["正面"] += 1
            else:
                sentiment["中性"] += 1

            # 提取讨论主题
            for topic in ["quality", "price", "comparison", "recommendation", "review", "experience"]:
                if topic in text:
                    topic_counts[topic] += 1

        # 主题映射
        topic_map = {
            "quality": "产品质量对比",
            "price": "性价比讨论",
            "comparison": "品牌推荐",
            "recommendation": "购买建议",
            "review": "使用体验分享",
            "experience": "售后吐槽",
        }

        topics = {}
        for eng, cnt in topic_counts.most_common(8):
            topics[topic_map.get(eng, eng)] = cnt

        demand_gaps = [
            {
                "frequency": count,
                "opportunity": topic,
                "description": f"Reddit 讨论中围绕“{topic}”出现 {count} 次有效信号",
                "action": "结合 Amazon 差评和竞品对标验证是否值得转化为产品改进点",
                "score": min(100, 60 + count),
            }
            for topic, count in list(topics.items())[:3]
        ]

        return {
            "total_posts": len(posts),
            "sentiment": sentiment,
            "top_topics": topics,
            "demand_gaps": demand_gaps,
            "raw_posts": posts,
        }

    # ================================================================
    # 5. Google Trends 数据 (NEW)
    # ================================================================

    def fetch_google_trends(self, keywords: List[str]) -> Dict:
        """Apify First 采集最近 6 个月 Google Trends 数据与截图引用。"""
        print("   ├─ 采集 Google Trends 数据...")
        client = self.apify_client
        actor_id = os.getenv("APIFY_GOOGLE_TRENDS_ACTOR_ID", "").strip()
        if not client or not actor_id:
            print("      ⚠️ Google Trends Actor 未配置，交由 Web Search/Web Fetch 层兜底")
            return {"keywords": {}, "timeline": [], "source": "unavailable", "fallback": "web_search"}
        try:
            import datetime
            kw_list = keywords[:5]
            trends_cfg = self.config.get("trends", {})
            months = max(1, int(trends_cfg.get("timeframe_months", 6)))
            end = datetime.date.today()
            start = end - datetime.timedelta(days=int(round(months * 30.4)))
            timeframe = f"{start:%Y-%m-%d} {end:%Y-%m-%d}"
            run_input = {
                "searchTerms": kw_list,
                "geo": self.market,
                "startDate": f"{start:%Y-%m-%d}",
                "endDate": f"{end:%Y-%m-%d}",
                "includeScreenshot": bool(trends_cfg.get("save_screenshot", True)),
            }
            run = client.actor(actor_id).call(run_input=run_input)
            dataset_id = _run_attr(run, "defaultDatasetId")
            items = list(client.dataset(dataset_id).iterate_items()) if dataset_id else []
            if not items:
                return {"keywords": {}, "timeline": [], "range": timeframe, "source": "unavailable", "fallback": "web_search"}

            timeline = []
            keywords_data = {kw: [] for kw in kw_list}
            points = items[0].get("interestOverTime") or items[0].get("timelineData") or items[0].get("timeline") or []
            for point in points:
                if not isinstance(point, dict):
                    continue
                timeline.append(str(point.get("date") or point.get("time") or point.get("formattedTime") or ""))
                values = point.get("values") or point.get("value") or []
                if not isinstance(values, list):
                    values = [values]
                for index, kw in enumerate(kw_list):
                    raw_value = values[index] if index < len(values) else ""
                    if isinstance(raw_value, dict):
                        raw_value = raw_value.get("value", raw_value.get("extractedValue", ""))
                    keywords_data[kw].append(raw_value)

            screenshot = items[0].get("screenshotUrl") or items[0].get("screenshot_url") or items[0].get("snapshotUrl") or ""
            screenshot_path = ""
            if screenshot and str(screenshot).startswith("http"):
                try:
                    from urllib.request import urlopen
                    screenshot_dir = Path(self.config.get("output_dir", "./output")) / "google_trends"
                    screenshot_dir.mkdir(parents=True, exist_ok=True)
                    safe_keyword = re.sub(r"[^A-Za-z0-9_-]+", "_", kw_list[0])[:50] or "trend"
                    target = screenshot_dir / f"{safe_keyword}_{end:%Y%m%d}.png"
                    with urlopen(screenshot, timeout=30) as response:
                        target.write_bytes(response.read())
                    screenshot_path = str(target.resolve())
                except Exception as exc:
                    print(f"      ⚠️ Google Trends 截图下载失败，保留远程 URL: {str(exc)[:100]}")
            return {
                "keywords": keywords_data,
                "timeline": timeline,
                "range": timeframe,
                "geo": self.market,
                "source": actor_id,
                "run_id": _run_attr(run, "id"),
                "dataset_id": dataset_id,
                "screenshot_url": screenshot,
                "screenshot_path": screenshot_path,
                "source_url": items[0].get("url", "https://trends.google.com/trends/"),
                "raw_items": items,
            }
        except Exception as e:
            print(f"      ⚠️ Google Trends Apify 采集失败，交由 Web Search/Web Fetch 兜底: {e}")
            return {"keywords": {}, "timeline": [], "source": "unavailable", "fallback": "web_search"}

    # ================================================================
    # 6. 1688 供应商数据
    # ================================================================

    def fetch_supplier_data(self, supplier_keywords: List[str]) -> Dict:
        """采集 1688 供应商报价数据
        
        v3.1: 1688 Actor 不可用，改用 Web Search 补充。
        """
        print("   ├─ 采集 1688 供应商数据...")
        client = self.apify_client
        
        # 尝试的 Actor IDs (按优先级)
        actor_candidates = [
            os.getenv("APIFY_1688_ACTOR_ID", "").strip(),
            "luzhiyu/1688-product-scraper",
            "easyapi/1688-product-scraper", 
            "logical_api/1688-scraper",
        ]
        actor_candidates = list(dict.fromkeys(actor for actor in actor_candidates if actor))

        if not client or not supplier_keywords:
            print("      ⚠️ 无 1688 Actor 可用，由 Web Search 补充")
            return {"suppliers": [], "total": 0, "_source": "unavailable"}

        all_suppliers = []
        for actor_id in actor_candidates:
            for kw in supplier_keywords[:1]:
                try:
                    run_input = {
                        "searchTerms": [kw],
                        "maxItems": 20,
                        "proxyConfiguration": {"useApifyProxy": True}
                    }
                    try:
                        run = client.actor(actor_id).start(
                            run_input=run_input,
                            wait_for_finish=120
                        )
                    except TypeError:
                        actor_call = client.actor(actor_id).start(run_input=run_input)
                        run = client.actor(actor_call["id"]).wait_for_finish(timeout_secs=120)
                    
                    if _run_attr(run, "status", "UNKNOWN") != "SUCCEEDED":
                        continue
                    
                    dataset_items = list(client.dataset(_run_attr(run, "defaultDatasetId")).iterate_items())
                    for item in dataset_items:
                        item_url = item.get("url", "")
                        if not item_url and kw:
                            item_url = f"https://www.1688.com/s?keywords={kw}"
                        all_suppliers.append({
                            "keyword": kw,
                            "title": item.get("title", ""),
                            "supplier": item.get("seller", ""),
                            "price_min": item.get("priceMin", 0),
                            "price_max": item.get("priceMax", 0),
                            "moq": item.get("minOrder", 0),
                            "location": item.get("province", ""),
                            # 来源链接：每条供应商对应 1688 搜索/商品页，便于人工核对报价
                            "source_url": item_url,
                        })
                    time.sleep(2)
                except Exception:
                    continue
            
            if all_suppliers:
                break  # 成功采集则退出循环

        if not all_suppliers:
            print("      ⚠️ 1688 无可用 Actor，由 Web Search 补充")
            return {"suppliers": [], "total": 0, "_source": "unavailable"}

        print(f"      ✅ 1688: {len(all_suppliers)} 个供应商")
        return {"suppliers": all_suppliers, "total": len(all_suppliers)}

    # ================================================================
    # 7. 市场报告数据
    # ================================================================

    def fetch_market_data(self, category: str) -> Dict:
        """采集市场数据
        
        v3.1: apify/web-scraper 需要账户审批，改用 Web Search 补充
        """
        print("   └─ 采集市场报告数据...")
        client = self.apify_client

        if not client:
            print("      ⚠️ 市场数据由 Web Search 补充")
            return {"category": category, "market_size": "", "cagr": 0, "source": "Web Search", "_source": "unavailable"}

        try:
            report_url = f"https://www.grandviewresearch.com/industry-analysis/{category.replace(' ', '-').lower()}-market"
            actor_id = os.getenv("APIFY_MARKET_REPORT_ACTOR_ID", "apify/web-scraper")

            run_input = {
                "startUrls": [{"url": report_url}],
                "maxRequestsPerCrawl": 1,
                "proxyConfiguration": {"useApifyProxy": True}
            }

            try:
                run = client.actor(actor_id).start(
                    run_input=run_input,
                    wait_for_finish=120
                )
            except TypeError:
                actor_call = client.actor(actor_id).start(run_input=run_input)
                run = client.actor(actor_call["id"]).wait_for_finish(timeout_secs=120)
            
            if _run_attr(run, "status", "UNKNOWN") == "SUCCEEDED":
                dataset_items = list(client.dataset(_run_attr(run, "defaultDatasetId")).iterate_items())
                if dataset_items:
                    text = dataset_items[0].get("text", "")
                    market_size_match = re.search(r'\$(\d+(?:\.\d+)?)\s*(billion|million|trillion)', text, re.IGNORECASE)
                    cagr_match = re.search(r'CAGR\s*(?:of\s*)?(\d+(?:\.\d+)?)\s*%', text, re.IGNORECASE)

                    result = {
                        "category": category,
                        "market_size": f"${market_size_match.group(1)} {market_size_match.group(2)}" if market_size_match else "",
                        "cagr": float(cagr_match.group(1)) / 100 if cagr_match else 0,
                        "source": f"Grand View Research (via {actor_id})",
                        "actor_id": actor_id,
                        "run_id": _run_attr(run, "id"),
                        "dataset_id": _run_attr(run, "defaultDatasetId"),
                        "source_url": report_url,
                    }
                    print(f"      ✅ 市场数据: {result.get('market_size', '')}")
                    return result
        except Exception as e:
            print(f"      ⚠️ 市场数据 Apify 采集失败: {str(e)[:120]}")

        print("      ⚠️ 市场数据由 Web Search 补充")
        return {"category": category, "market_size": "", "cagr": 0, "source": "Web Search", "_source": "unavailable"}

    # ================================================================
    # 综合采集入口
    # ================================================================

    def collect_all(self, keywords: List[str], asins: List[str] = None,
                    supplier_keywords: List[str] = None,
                    subreddits: List[str] = None) -> dict:
        """
        一站式全量数据采集

        Returns:
            完整数据字典，包含所有数据源
        """
        print("\n📊 开始全量数据采集...")
        print(f"   市场: {self.market} | 关键词: {keywords}")
        print("-" * 50)

        data = {}

        # Amazon 商品
        data["amazon"] = self.fetch_amazon_products(keywords)

        # 差评分析
        if asins:
            data["review_analysis"] = self.fetch_negative_reviews(asins)
        else:
            # 从已采集的商品中提取 ASIN
            collected_asins = [p.get("asin") for p in data["amazon"] if p.get("asin")]
            data["review_analysis"] = self.fetch_negative_reviews(collected_asins[:20])

        # TikTok
        data["tiktok"] = self.fetch_tiktok_data(keywords)

        # TikTok Shop 销量数据（品类销量/GMV/价格/店铺）
        data["tiktok_shop"] = self.fetch_tiktok_shop_sales(keywords)

        # Reddit
        data["reddit"] = self.fetch_reddit_data(keywords, subreddits)

        # Google Trends
        data["google_trends"] = self.fetch_google_trends(keywords)

        # 1688 供应商
        data["supplier"] = self.fetch_supplier_data(supplier_keywords or [])

        # 市场报告
        data["market"] = self.fetch_market_data(
            self.config.get("category", keywords[0]) if isinstance(keywords[0], str) else keywords[0]
        )

        print(f"\n✅ 全量采集完成!")
        print(f"   Amazon: {len(data['amazon'])} 商品")
        print(f"   差评: {data['review_analysis'].get('total_reviews', 0)} 条")
        print(f"   TikTok: {len(data['tiktok'])} 标签")
        print(f"   TikTok Shop: {len(data.get('tiktok_shop', []))} 个商品销量")
        print(f"   Reddit: {data['reddit'].get('total_posts', 0)} 帖子")
        print(f"   1688: {data['supplier'].get('total', 0)} 供应商")

        return data

# ================================================================
# 便捷函数
# ================================================================

def collect_sample_data(category: str, market: str = "US") -> dict:
    """生成空数据结构用于管道测试，不填充示例或 mock 数据。"""
    keywords = [category]

    data = {
        "metadata": {
            "market": market,
            "category": category,
            "keywords": keywords,
            "generated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        },
        "amazon": [],
        "tiktok": [],
        "review_analysis": {"total_reviews": 0, "pain_points": [], "pain_categories": {}, "top_negative_keywords": []},
        "reddit": {"total_posts": 0, "sentiment": {}, "top_topics": {}, "demand_gaps": []},
        "google_trends": {"keywords": {}, "timeline": [], "source": "unavailable"},
        "supplier": {"suppliers": [], "total": 0, "_source": "unavailable"},
        "market": {"category": category, "market_size": "", "cagr": 0, "source": "unavailable"},
        "analysis": {},
    }

    return data


if __name__ == "__main__":
    # 测试空数据结构生成
    sample = collect_sample_data("kids supplements", "US")
    print(f"✅ 空数据结构生成完成")
    print(f"   Amazon: {len(sample['amazon'])} 商品")
    print(f"   TikTok: {len(sample['tiktok'])} 标签")
    print(f"   差评: {sample['review_analysis']['total_reviews']} 条")
    print(f"   Reddit: {sample['reddit']['total_posts']} 帖子")
