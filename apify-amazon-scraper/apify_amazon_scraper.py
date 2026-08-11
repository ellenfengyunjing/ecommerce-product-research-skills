#!/usr/bin/env python3
"""
Apify Amazon Scraper - 采集亚马逊商品数据
使用方法: python apify_amazon_scraper.py <关键词> <数量> [输出目录]

依赖安装: pip install apify-client pandas openpyxl
"""

import sys
import os
from datetime import datetime

try:
    from apify_client import ApifyClient
except ImportError:
    print("请先安装 apify-client: pip install apify-client")
    sys.exit(1)

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("请先安装依赖: pip install pandas openpyxl")
    sys.exit(1)


def get_apify_token():
    """从环境变量或配置文件获取API Token"""
    token = os.environ.get("APIFY_API_TOKEN")
    if token:
        return token

    config_path = os.path.expanduser("~/.workbuddy/apify_config.json")
    if os.path.exists(config_path):
        import json
        with open(config_path) as f:
            config = json.load(f)
            return config.get("api_token", "")

    return input("请输入 Apify API Token: ").strip()


def create_search_url(keyword, sort="sales"):
    """创建Amazon搜索URL"""
    import urllib.parse
    encoded = urllib.parse.quote(keyword)
    # sales = 按销量排序, review_rank = 按评论排序
    return f"https://www.amazon.com/s?k={encoded}&s={sort}"


def scrape_amazon(client, search_url, max_items=50):
    """调用Apify Actor抓取数据"""
    run_input = {
        "searchUrls": [search_url],
        "maxItems": max_items,
        "scrapeVariants": True,
        "scrapeFromCategory": True,
        "includeFields": [
            "title", "url", "asin", "brand", "price",
            "stars", "reviewsCount", "bestsellerRanks",
            "images", "description", "variantAsins",
            "features", "dateFirstAvailable"
        ]
    }

    print(f"启动Apify Actor抓取任务...")
    actor_call = client.actor("junglee/free-amazon-product-scraper").start(run_input=run_input)

    print(f"Actor运行ID: {actor_call['id']}, 等待完成...")
    client.actor(actor_call["id"]).wait_for_finish()

    print(f"获取数据集...")
    dataset_items = client.dataset(actor_call["defaultDatasetId"]).list_items().items

    return dataset_items


def transform_item(item):
    """转换数据字段"""
    # 提取主图
    main_image = ""
    if item.get("images"):
        if isinstance(item["images"], list):
            main_image = item["images"][0].get("link", "") if isinstance(item["images"][0], dict) else str(item["images"][0])
        elif isinstance(item["images"], dict):
            main_image = item["images"].get("link", "")

    # 提取Buy Box价格
    buy_box_price = ""
    price_data = item.get("price", {})
    if isinstance(price_data, dict):
        buy_box_price = f"{price_data.get('currency', '$')}{price_data.get('value', '')}"
    elif price_data:
        buy_box_price = str(price_data)

    # 提取BSR排名
    bsr_list = item.get("bestsellerRanks", [])
    bsr_text = ""
    if isinstance(bsr_list, list):
        bsr_text = " | ".join([f"#{r.get('rank', '')} in {r.get('category', '')}" for r in bsr_list[:3]])
    elif bsr_list:
        bsr_text = str(bsr_list)

    # 提取分类
    category = ""
    bread_crumbs = item.get("breadCrumbs", "")
    if bread_crumbs:
        category = " > ".join(bread_crumbs) if isinstance(bread_crumbs, list) else bread_crumbs

    # 提取变体数量
    variant_count = 0
    variants = item.get("variantAsins", [])
    if isinstance(variants, list):
        variant_count = len(variants)

    # 提取上架日期
    launch_date = item.get("dateFirstAvailable", "")
    if isinstance(launch_date, dict):
        launch_date = launch_date.get("value", "")

    return {
        "标题": item.get("title", ""),
        "品牌": item.get("brand", ""),
        "分类": category,
        "首图URL": main_image,
        "上架日期": launch_date,
        "BSR排名": bsr_text,
        "预估月销量": "",  # 需要从搜索页提取
        "价格历史": "",     # 需要单独接口
        "Buy Box价格": buy_box_price,
        "评论数": item.get("reviewsCount", 0),
        "评分": item.get("stars", 0.0),
        "变体数量": variant_count
    }


def save_to_excel(data, output_path):
    """保存数据到Excel并格式化"""
    df = pd.DataFrame(data)

    # 保存
    df.to_excel(output_path, index=False, engine='openpyxl')

    # 格式化
    wb = load_workbook(output_path)
    ws = wb.active

    # 表头样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # 自动列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    # 冻结首行
    ws.freeze_panes = "A2"

    wb.save(output_path)
    return output_path


def main():
    # 解析参数
    if len(sys.argv) < 3:
        print("用法: python apify_amazon_scraper.py <关键词> <数量> [输出目录]")
        print("示例: python apify_amazon_scraper.py children's health products 50")
        sys.exit(1)

    keyword = sys.argv[1]
    max_items = int(sys.argv[2])
    output_dir = sys.argv[3] if len(sys.argv) > 3 else None

    # 设置输出目录
    if not output_dir:
        output_dir = os.path.join(os.path.expanduser("~"), "Desktop")
    os.makedirs(output_dir, exist_ok=True)

    # 生成文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_keyword = keyword.replace(" ", "_").replace("'", "")[:20]
    excel_file = os.path.join(output_dir, f"amazon_{safe_keyword}_{timestamp}.xlsx")

    # 获取Token
    api_token = get_apify_token()
    if not api_token:
        print("错误: 请设置APIFY_API_TOKEN环境变量或创建配置文件")
        sys.exit(1)

    # 初始化客户端
    client = ApifyClient(api_token)

    # 创建搜索URL
    search_url = create_search_url(keyword)

    print(f"\n关键词: {keyword}")
    print(f"最大数量: {max_items}")
    print(f"搜索URL: {search_url}\n")

    # 抓取数据
    raw_data = scrape_amazon(client, search_url, max_items)

    print(f"获取到 {len(raw_data)} 条原始数据")

    # 转换数据
    transformed_data = [transform_item(item) for item in raw_data[:max_items]]

    # 保存Excel
    excel_path = save_to_excel(transformed_data, excel_file)

    print(f"\n✅ 数据已保存到: {excel_path}")
    print(f"共 {len(transformed_data)} 条商品记录")


if __name__ == "__main__":
    main()
