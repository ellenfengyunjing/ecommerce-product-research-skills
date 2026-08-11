"""
Excel Exporter for Amazon Product Researcher (产物一：采集数据 Excel 工作簿)
==========================================================================

将 7 大数据源采集到的原始数据，统一导出为一个**全面、准确、可溯源**的 Excel 工作簿。

设计原则（对齐用户要求）：
  1. Amazon 来自 amazon-product-scraper 优先、Apify 补缺；其他平台来自 Apify 优先。
  2. 每个工作表的每一行都带有「来源链接」列，方便人工逐条核对真实性。
  3. 平台额度：Amazon 商品默认 100、硬下限 50；TikTok 原始视频 ≥50；
     1688 有效报价 ≥20；Reddit 可追溯帖子/评论 ≥20；Google Trends 保存半年数据与截图。
  4. 标准化工作表外，额外输出「原始字段明细」长表和 Data_Lineage，避免字段丢失。
  5. 采集失败/缺失的数据不编造，工作表仅保留表头或标注「本轮未采集」。

用法：
    python excel_exporter.py --input raw_data.json --output report.xlsx
    python excel_exporter.py --input-dir ./output/2026-xx_xx_full_api --output report.xlsx

依赖：openpyxl>=3.0.10
"""

import argparse
import json
import os
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write("⚠️ 缺少 openpyxl，请先安装：pip install openpyxl\n")
    raise


# ---------------------------------------------------------------------------
# 样式
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1A73E8")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color="1A73E8", bold=True, size=14)
NOTE_FONT = Font(color="8B949E", italic=True, size=10)
URL_FONT = Font(color="0563C1", underline="single", size=10)
THIN = Side(style="thin", color="D0D7DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(vertical="top", wrap_text=True)


# ---------------------------------------------------------------------------
# 数据加载
# ---------------------------------------------------------------------------

def load_data(input_path: str) -> Dict[str, Any]:
    p = Path(input_path)
    if p.is_dir():
        return _load_from_dir(p)
    return json.loads(p.read_text(encoding="utf-8"))


def _load_from_dir(d: Path) -> Dict[str, Any]:
    """从分阶段的 raw JSON 文件合并成一个统一的 combined 结构。"""
    merged: Dict[str, Any] = {}
    stage_map = {
        "raw_after_amazon.json": "amazon",
        "raw_after_reviews.json": "reviews",
        "raw_after_tiktok.json": "tiktok",
        "raw_after_reddit.json": "reddit",
        "raw_after_suppliers.json": "suppliers_stage",
        "raw_after_trends.json": "google_trends",
    }
    for fname, key in stage_map.items():
        fp = d / fname
        if fp.exists():
            try:
                merged[key] = json.loads(fp.read_text(encoding="utf-8"))
            except Exception:
                pass
    # 统一为 combined 结构
    out: Dict[str, Any] = {
        "metadata": merged.get("amazon", {}).get("metadata", {}) if isinstance(merged.get("amazon"), dict) else {},
    }
    if "amazon" in merged and isinstance(merged["amazon"], dict) and "amazon_products" in merged["amazon"]:
        out["amazon"] = merged["amazon"]["amazon_products"]
    if "reviews" in merged and isinstance(merged["reviews"], dict):
        rev = merged["reviews"].get("reviews", {})
        out["review_analysis"] = {
            "raw_reviews": rev.get("raw_reviews", []),
            "total_reviews": rev.get("analysis", {}).get("total_reviews", 0),
        }
    if "tiktok" in merged and isinstance(merged["tiktok"], dict):
        out["tiktok"] = merged["tiktok"].get("tiktok", [])
    if "reddit" in merged:
        rd = merged["reddit"]
        if isinstance(rd, dict) and "reddit" in rd:
            out["reddit"] = {"raw_posts": rd.get("reddit", [])}
        elif isinstance(rd, list):
            out["reddit"] = {"raw_posts": rd}
    if "suppliers_stage" in merged and isinstance(merged["suppliers_stage"], dict):
        out["supplier"] = {"suppliers": merged["suppliers_stage"].get("supplier_pages", [])}
    if "google_trends" in merged and isinstance(merged["google_trends"], dict):
        gt = merged["google_trends"].get("google_trends", merged["google_trends"])
        out["google_trends"] = gt
    out["_lineage"] = merged.get("amazon", {}).get("lineage", {}) if isinstance(merged.get("amazon"), dict) else {}
    return out


# ---------------------------------------------------------------------------
# 归一化：把各平台原始结构整理成「行字典列表」
# ---------------------------------------------------------------------------

NEG_REDDIT_KW = [
    "hate", "worst", "terrible", "avoid", "waste", "disappointed",
    "regret", "broke", "returned", "scam", "fake", "broken", "useless",
    "don't buy", "dont buy", "not worth", "overpriced", "garbage", "awful",
]


def _first(item: Dict[str, Any], *keys: str, default: Any = "") -> Any:
    for key in keys:
        value = item.get(key)
        if value not in (None, "", [], {}):
            return value
    return default


def _cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return value


def _provenance(item: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "采集方式": _first(item, "collection_method", "source", "collector"),
        "Actor/脚本": _first(item, "actor_id", "actor", "script", "source"),
        "Run ID": _first(item, "run_id", "runId"),
        "Dataset ID": _first(item, "dataset_id", "datasetId", "defaultDatasetId"),
        "采集时间": _first(item, "collected_at", "collection_date", "scraped_at", "timestamp"),
    }


def _norm_amazon(data: Dict[str, Any]) -> List[Dict]:
    products = data.get("amazon") or []
    if not isinstance(products, list):
        return []
    rows = []
    for p in products:
        if not isinstance(p, dict):
            continue
        row = {
            "ASIN": p.get("asin", ""),
            "商品标题": p.get("title", ""),
            "品牌": p.get("brand", ""),
            "类目": _cell_value(_first(p, "category", "categories", "breadCrumbs")),
            "价格": _cell_value(_first(p, "price", "buy_box_price", "buyBoxPrice")),
            "币种": _first(p, "currency", "currencyCode"),
            "Buy Box价": _cell_value(_first(p, "buy_box_price", "buyBoxPrice")),
            "评分": _first(p, "rating", "stars", "review_rating"),
            "评论数": _first(p, "review_count", "reviewsCount", "reviewCount"),
            "BSR排名": _cell_value(_first(p, "bsr", "bestsellerRanks", "sales_rank_BSR")),
            "预估月销量": _first(p, "estimated_monthly_sales", "estimatedMonthlySales", "monthly_sales", "sales"),
            "上架日期": _cell_value(_first(p, "launch_date", "dateFirstAvailable")),
            "变体数": _first(p, "variation_count", "variationCount"),
            "五点描述": _cell_value(_first(p, "bullets", "features", "bulletPoints")),
            "规格参数": _cell_value(_first(p, "specifications", "specs", "technicalDetails")),
            "首图URL": _cell_value(_first(p, "image", "main_image", "images")),
            "关键词": p.get("keyword", ""),
            "市场": p.get("market", ""),
            "来源链接": p.get("source_url", ""),
        }
        row.update(_provenance(p))
        rows.append(row)
    return rows


def _norm_reviews(data: Dict[str, Any]) -> List[Dict]:
    rev = data.get("review_analysis") or {}
    raw = rev.get("raw_reviews") if isinstance(rev, dict) else None
    if not raw and isinstance(data.get("reviews"), dict):
        raw = data["reviews"].get("raw_reviews", [])
    if not raw:
        return []
    rows = []
    for r in raw:
        if not isinstance(r, dict):
            continue
        row = {
            "Review ID": _first(r, "review_id", "reviewId", "id"),
            "ASIN": r.get("asin", ""),
            "评分": r.get("rating", ""),
            "评论标题": r.get("title", ""),
            "评论内容": r.get("text", ""),
            "日期": r.get("date", ""),
            "已验证购买": r.get("verified", ""),
            "有用票数": _first(r, "helpful_votes", "helpfulVotes", "helpful"),
            "变体": _cell_value(_first(r, "variant", "variation")),
            "来源链接": _first(r, "source_url", "review_url", "url"),
        }
        row.update(_provenance(r))
        rows.append(row)
    return rows


def _norm_tiktok(data: Dict[str, Any]) -> List[Dict]:
    items = data.get("tiktok") or []
    if isinstance(items, dict):
        items = items.get("videos") or items.get("items") or items.get("data") or []
    if not isinstance(items, list):
        return []
    rows = []
    for t in items:
        if not isinstance(t, dict):
            continue
        row = {
            "Video ID": _first(t, "video_id", "videoId", "id"),
            "作者": _first(t, "author", "profile_name", "author_name"),
            "作者URL": _first(t, "author_url", "profile_url"),
            "文案": _first(t, "description", "text"),
            "标签": _cell_value(_first(t, "hashtags", "hashtag")),
            "时长(秒)": _first(t, "duration"),
            "播放量": _first(t, "views", "plays", "playCount", "sample_total_views"),
            "点赞数": _first(t, "likes", "diggCount", "sample_total_likes"),
            "评论数": _first(t, "comments", "comment_count", "commentCount", "sample_total_comments"),
            "分享数": _first(t, "shares", "shareCount"),
            "收藏数": _first(t, "saves", "collectCount"),
            "发布时间": _first(t, "create_time", "createTimeISO", "createTime"),
            "音乐": _first(t, "music", "music_name"),
            "关键词": t.get("keyword", ""),
            "来源链接": _first(t, "source_url", "video_url", "webVideoUrl", "url"),
        }
        row.update(_provenance(t))
        rows.append(row)
    return rows


def _norm_tiktok_shop(data: Dict[str, Any]) -> List[Dict]:
    items = data.get("tiktok_shop") or []
    if not isinstance(items, list):
        return []
    rows = []
    for t in items:
        if not isinstance(t, dict):
            continue
        row = {
            "商品ID": _first(t, "product_id", "productId", "id"),
            "关键词": t.get("keyword", ""),
            "商品标题": t.get("product_title", ""),
            "类目": _first(t, "category", "category_name"),
            "价格": t.get("price", ""),
            "币种": _first(t, "currency", "currencyCode"),
            "销量": t.get("sales_volume", ""),
            "GMV": t.get("gmv", ""),
            "店铺": t.get("shop_name", ""),
            "评分": t.get("rating", ""),
            "评论数": _first(t, "review_count", "reviewsCount"),
            "佣金": _first(t, "commission", "commission_rate"),
            "来源链接": _first(t, "source_url", "product_url", "url"),
        }
        row.update(_provenance(t))
        rows.append(row)
    return rows


def _norm_reddit_posts(data: Dict[str, Any]) -> List[Dict]:
    rd = data.get("reddit") or {}
    raw = rd.get("raw_posts") if isinstance(rd, dict) else None
    if not raw and isinstance(rd, list):
        raw = rd
    if not raw:
        return []
    rows = []
    for p in raw:
        if not isinstance(p, dict):
            continue
        row = {
            "Post/Comment ID": _first(p, "post_id", "comment_id", "id"),
            "子版块": p.get("subreddit", ""),
            "作者": _first(p, "author", "username"),
            "标题": p.get("title", ""),
            "正文": p.get("text", p.get("body", "")),
            "类型": _first(p, "type", "kind"),
            "情绪": _first(p, "sentiment"),
            "点赞": p.get("score", ""),
            "评论数": p.get("num_comments", p.get("numberOfComments", "")),
            "发布时间": _first(p, "created_at", "created_utc", "createdAt"),
            "关键词": _first(p, "keyword", "query"),
            "来源链接": _first(p, "url", "source_url", "permalink"),
        }
        row.update(_provenance(p))
        rows.append(row)
    return rows


def _norm_reddit_negative(data: Dict[str, Any]) -> List[Dict]:
    posts = _norm_reddit_posts(data)
    neg = []
    for row in posts:
        text = f"{row.get('标题','')} {row.get('正文','')}".lower()
        if any(k in text for k in NEG_REDDIT_KW):
            row["负面命中词"] = ", ".join(k for k in NEG_REDDIT_KW if k in text)
            neg.append(row)
    # 若无法判定负面（文本缺失），退回全部帖子由人工筛选
    return neg if neg else posts


def _norm_suppliers(data: Dict[str, Any]) -> List[Dict]:
    sup = data.get("supplier") or {}
    sup_list = sup.get("suppliers") if isinstance(sup, dict) else None
    if not sup_list and isinstance(sup, dict):
        sup_list = sup.get("supplier_pages", [])
    if not sup_list:
        return []
    rows = []
    for s in sup_list:
        if not isinstance(s, dict):
            continue
        # 兼容 Apify 供应商结构 与 1688/Alibaba 搜索页结构
        price_min = s.get("price_min", s.get("priceMin", ""))
        price_max = s.get("price_max", s.get("priceMax", ""))
        moq_samples = s.get("moq_samples", [])
        moq = s.get("moq", s.get("minOrder", moq_samples[0] if moq_samples else ""))
        price_samples = s.get("price_samples_usd", [])
        title = s.get("title", "")
        if not title and price_samples:
            title = f"价格样本: {price_samples}"
        row = {
            "Offer ID": _first(s, "offer_id", "offerId", "id"),
            "关键词": s.get("keyword", s.get("query", "")),
            "商品/标题": title,
            "供应商": s.get("supplier", s.get("seller", "")),
            "阶梯价/价格区间": _cell_value(_first(s, "tier_prices", "priceTiers", default=(f"{price_min}-{price_max}" if price_min != "" or price_max != "" else price_samples))),
            "币种": _first(s, "currency", default="CNY"),
            "MOQ": moq,
            "单位": _first(s, "unit", "salesUnit"),
            "评分": _first(s, "rating", "sellerRating"),
            "复购率": _first(s, "repurchase_rate", "repurchaseRate"),
            "成交量": _first(s, "sales", "transactions", "sold"),
            "响应率": _first(s, "response_rate", "responseRate"),
            "地区": s.get("location", s.get("province", "")),
            "经营年限": _first(s, "years", "yearsInBusiness"),
            "认证": _cell_value(_first(s, "certifications", "certificates")),
            "OEM/ODM": _first(s, "oem_odm", "oem", "odm"),
            "交期": _first(s, "lead_time", "leadTime"),
            "来源链接": _first(s, "source_url", "url", "productUrl"),
        }
        row.update(_provenance(s))
        rows.append(row)
    return rows


def _norm_market(data: Dict[str, Any]) -> List[Dict]:
    m = data.get("market") or {}
    if not isinstance(m, dict):
        return []
    row = {
        "报告名": _first(m, "report_name", "title"),
        "发布机构": _first(m, "publisher", "organization", "source"),
        "类目": m.get("category", ""),
        "地区": _first(m, "region", "market"),
        "基准年": _first(m, "base_year", "baseYear"),
        "市场规模": m.get("market_size", ""),
        "预测年": _first(m, "forecast_year", "forecastYear"),
        "预测值": _first(m, "forecast_value", "forecastValue"),
        "CAGR": m.get("cagr", ""),
        "发布时间": _first(m, "published_at", "publication_date", "year"),
        "摘要": _first(m, "summary", "excerpt"),
        "来源": m.get("source", ""),
        "来源链接": m.get("source_url", m.get("url", "")),
    }
    row.update(_provenance(m))
    return [row]


def _platform_records(data: Dict[str, Any]):
    """迭代所有原始记录，供「原始字段明细」工作表无损展开。"""
    review_data = data.get("review_analysis") or data.get("reviews") or {}
    reddit_data = data.get("reddit") or {}
    supplier_data = data.get("supplier") or {}
    tiktok_data = data.get("tiktok") or []
    if isinstance(tiktok_data, dict):
        tiktok_data = tiktok_data.get("videos") or tiktok_data.get("items") or tiktok_data.get("data") or []

    groups = {
        "Amazon": data.get("amazon") or [],
        "Amazon Reviews": review_data.get("raw_reviews", []) if isinstance(review_data, dict) else [],
        "TikTok": tiktok_data,
        "TikTok Shop": data.get("tiktok_shop") or [],
        "Reddit": reddit_data.get("raw_posts", []) if isinstance(reddit_data, dict) else reddit_data,
        "1688": supplier_data.get("suppliers", supplier_data.get("supplier_pages", [])) if isinstance(supplier_data, dict) else supplier_data,
        "Market Report": [data.get("market")] if isinstance(data.get("market"), dict) and data.get("market") else [],
        "Google Trends": [data.get("google_trends")] if isinstance(data.get("google_trends"), dict) and data.get("google_trends") else [],
    }
    for platform, records in groups.items():
        if not isinstance(records, list):
            continue
        for index, record in enumerate(records, 1):
            if isinstance(record, dict):
                yield platform, index, record


def _flatten(value: Any, prefix: str = ""):
    if isinstance(value, dict):
        if not value:
            yield prefix, "{}"
        for key, child in value.items():
            child_prefix = f"{prefix}.{key}" if prefix else str(key)
            yield from _flatten(child, child_prefix)
    elif isinstance(value, list):
        if not value:
            yield prefix, "[]"
        for index, child in enumerate(value):
            yield from _flatten(child, f"{prefix}[{index}]")
    else:
        yield prefix, value


def _raw_field_rows(data: Dict[str, Any]) -> List[Dict]:
    rows = []
    for platform, index, record in _platform_records(data):
        source_url = _first(record, "source_url", "url", "webVideoUrl", "video_url", "productUrl")
        for field_path, value in _flatten(record):
            rows.append({
                "平台": platform,
                "记录索引": index,
                "字段路径": field_path,
                "原始值": _cell_value(value),
                "来源链接": source_url,
            })
    return rows


def _lineage_rows(data: Dict[str, Any]) -> List[Dict]:
    lineage = data.get("_lineage") or data.get("lineage") or (data.get("metadata") or {}).get("data_lineage") or {}
    rows = []
    if not isinstance(lineage, dict):
        return rows
    for platform, entry in lineage.items():
        item = entry if isinstance(entry, dict) else {"source": entry}
        rows.append({
            "平台/数据集": platform,
            "采集方式": _first(item, "collection_method", "method", "source"),
            "Actor/脚本": _first(item, "actor_id", "actor", "script", "source"),
            "Run ID": _first(item, "run_id", "runId"),
            "Dataset ID": _first(item, "dataset_id", "datasetId", "defaultDatasetId"),
            "查询/范围": _cell_value(_first(item, "query", "keywords", "scope", "asins")),
            "请求数": _first(item, "requested_count", "requested"),
            "有效数": _first(item, "valid_count", "count", "sample_size"),
            "去重数": _first(item, "deduplicated_count", "dedup_count"),
            "采集时间": _first(item, "time", "collected_at", "collection_date", "timestamp"),
            "状态": _first(item, "status"),
            "补采/降级原因": _first(item, "fallback_reason", "supplement_reason", "reason"),
            "来源链接": _first(item, "source_url", "dataset_url", "actor_url", "url"),
        })
    return rows


# ---------------------------------------------------------------------------
# 写入工作表
# ---------------------------------------------------------------------------

def write_sheet(wb: Workbook, title: str, columns: List[str], rows: List[Dict],
                url_col: Optional[str] = None, note: str = "") -> None:
    ws = wb.create_sheet(title=title)
    ws.append(columns)
    for c in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        cell.border = BORDER

    if not rows:
        ws.append([note or "（本轮未采集 / 无数据）"] + [""] * (len(columns) - 1))
        ws.cell(row=2, column=1).font = NOTE_FONT
    else:
        for r in rows:
            ws.append([r.get(col, "") for col in columns])

        # 来源链接列 -> 超链接
        if url_col and url_col in columns:
            url_idx = columns.index(url_col) + 1
            for ridx in range(2, len(rows) + 2):
                val = ws.cell(row=ridx, column=url_idx).value
                if val and isinstance(val, str) and val.startswith("http"):
                    ws.cell(row=ridx, column=url_idx).hyperlink = val
                    ws.cell(row=ridx, column=url_idx).font = URL_FONT

    # 边框 + 自动列宽 + 冻结首行
    for ridx in range(2, len(rows) + 2 if rows else 3):
        for cidx in range(1, len(columns) + 1):
            cell = ws.cell(row=ridx, column=cidx)
            cell.border = BORDER
            cell.alignment = WRAP_TOP

    for c in range(1, len(columns) + 1):
        letter = get_column_letter(c)
        max_len = len(str(columns[c - 1]))
        for ridx in range(2, min(len(rows) + 2, 200) if rows else 3):
            v = ws.cell(row=ridx, column=c).value
            if v:
                max_len = max(max_len, min(len(str(v)), 80))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 60)
    ws.freeze_panes = "A2"


def write_overview(wb: Workbook, data: Dict[str, Any]) -> None:
    ws = wb.create_sheet("采集总览", 0)
    meta = data.get("metadata", {})
    ws["A1"] = "Amazon 选品调研 · 采集数据总览（产物一）"
    ws["A1"].font = TITLE_FONT
    ws.append([])
    info = [
        ("市场", meta.get("market", "")),
        ("类目", meta.get("category", "")),
        ("关键词", ", ".join(meta.get("keywords", []) or [])),
        ("采集时间", meta.get("generated_at", "")),
        ("Amazon 商品数", len(_norm_amazon(data))),
        ("Amazon 差评数", len(_norm_reviews(data))),
        ("TikTok 原始视频数(目标≥50)", len(_norm_tiktok(data))),
        ("TikTok Shop 销量条目", len(_norm_tiktok_shop(data))),
        ("Reddit 帖子数", len(_norm_reddit_posts(data))),
        ("Reddit 差评数(≥目标20)", len(_norm_reddit_negative(data))),
        ("1688 供应商数", len(_norm_suppliers(data))),
        ("Google Trends 关键词数", len((data.get("google_trends") or {}).get("keywords", {}) or {})),
        ("市场报告", "有" if _norm_market(data) else "无"),
    ]
    for k, v in info:
        ws.append([k, v])
    # 采集方式 lineage
    lineage = data.get("_lineage") or data.get("lineage") or meta.get("data_lineage") or {}
    if lineage:
        ws.append([])
        ws.append(["采集方式 (Data Lineage)", ""])
        for k, v in lineage.items():
            ws.append([f"  {k}", str(v)])
    for r in range(3, ws.max_row + 1):
        ws.cell(row=r, column=1).font = Font(bold=True)
        for c in (1, 2):
            ws.cell(row=r, column=c).border = BORDER
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 70


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------

def build_workbook(data: Dict[str, Any]) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认空表

    write_overview(wb, data)

    def add(title: str, rows: List[Dict], fallback_columns: List[str], note: str = "") -> None:
        columns = list(rows[0].keys()) if rows else fallback_columns
        write_sheet(wb, title, columns, rows, url_col="来源链接" if "来源链接" in columns else None, note=note)

    add("Amazon_商品", _norm_amazon(data), ["ASIN", "商品标题", "品牌", "类目", "价格", "评分", "评论数", "BSR排名", "来源链接"])
    add("Amazon_差评", _norm_reviews(data), ["Review ID", "ASIN", "评分", "评论标题", "评论内容", "日期", "已验证购买", "来源链接"])
    add("TikTok_流量", _norm_tiktok(data), ["Video ID", "作者", "文案", "标签", "播放量", "点赞数", "评论数", "分享数", "来源链接"])
    add("TikTok_Shop_销量", _norm_tiktok_shop(data), ["商品ID", "关键词", "商品标题", "价格", "销量", "GMV", "店铺", "来源链接"])
    add("Reddit_讨论", _norm_reddit_posts(data), ["Post/Comment ID", "子版块", "标题", "正文", "点赞", "评论数", "来源链接"])
    add("Reddit_差评", _norm_reddit_negative(data), ["Post/Comment ID", "子版块", "标题", "正文", "负面命中词", "来源链接"])
    add("1688_供应商", _norm_suppliers(data), ["Offer ID", "关键词", "商品/标题", "供应商", "阶梯价/价格区间", "MOQ", "来源链接"])

    # Google Trends 单独处理（时间×关键词矩阵）
    _write_trends(wb, data)

    add("市场报告", _norm_market(data), ["报告名", "发布机构", "类目", "市场规模", "CAGR", "来源链接"])
    add("Data_Lineage", _lineage_rows(data), ["平台/数据集", "采集方式", "Actor/脚本", "Run ID", "Dataset ID", "有效数", "采集时间", "状态", "来源链接"])
    add("原始字段明细", _raw_field_rows(data), ["平台", "记录索引", "字段路径", "原始值", "来源链接"])

    return wb


def _write_trends(wb: Workbook, data: Dict[str, Any]) -> None:
    gt = data.get("google_trends") or {}
    ws = wb.create_sheet("Google_Trends")
    ws.append(["Google Trends 关键词流量（最近 6 个月）"])
    ws["A1"].font = TITLE_FONT
    metadata = [
        ("时间范围", _first(gt, "range", "timeframe")),
        ("地区", _first(gt, "geo", "region")),
        ("采集方式", _first(gt, "source", "collector", "actor_id")),
        ("Run ID", _first(gt, "run_id", "runId")),
        ("Dataset ID", _first(gt, "dataset_id", "datasetId")),
        ("趋势截图", _first(gt, "screenshot_path", "screenshot_url", "screenshot")),
        ("来源链接", _first(gt, "source_url", "url")),
    ]
    for label, value in metadata:
        if value not in (None, "", [], {}):
            ws.append([label, _cell_value(value)])
            cell = ws.cell(row=ws.max_row, column=2)
            if isinstance(cell.value, str) and (cell.value.startswith("http") or Path(cell.value).exists()):
                cell.hyperlink = cell.value
                cell.font = URL_FONT
    keywords = gt.get("keywords", {}) or {}
    timeline = gt.get("timeline", []) or []
    if not keywords or not timeline:
        ws.append(["（本轮未采集到 Google Trends / 无数据）"])
        ws.cell(row=ws.max_row, column=1).font = NOTE_FONT
        return
    header = ["日期/周"] + list(keywords.keys())
    ws.append(header)
    header_row = ws.max_row
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
    for i, month in enumerate(timeline):
        row = [month] + [keywords[k][i] if i < len(keywords[k]) else "" for k in keywords.keys()]
        ws.append(row)
    for c in range(1, len(header) + 1):
        letter = get_column_letter(c)
        ws.column_dimensions[letter].width = 22
    ws.freeze_panes = f"B{header_row + 1}"


def main():
    ap = argparse.ArgumentParser(description="Export collected raw data to a comprehensive Excel workbook.")
    ap.add_argument("--input", help="combined raw_data.json path")
    ap.add_argument("--input-dir", help="directory containing stage raw JSON files")
    ap.add_argument("--output", required=True, help="output .xlsx path")
    args = ap.parse_args()

    if args.input_dir:
        data = load_data(args.input_dir)
    elif args.input:
        data = load_data(args.input)
    else:
        sys.stderr.write("⚠️ 必须指定 --input 或 --input-dir\n")
        sys.exit(2)

    wb = build_workbook(data)
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"✅ Excel 工作簿已生成: {out}")
    print(f"   工作表: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
