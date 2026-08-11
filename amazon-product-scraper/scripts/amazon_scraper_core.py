#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
亚马逊商品采集技能 - 核心采集脚本 v6.3
功能：搜索 Amazon.com Best Seller 商品，采集数据并导出 Excel + 下载首图

v6.3 升级项：
1. 【P0 价格精准修复】新增 corePriceDisplay_desktop / priceToPay 等精准容器选择器，
     在 Tier 0b 之前增加专属层，避免误命中划线价/配件价/分期价
2. 【P1 Tier 3 兜底优化】不再取"最低价"，改为位置加权+中位数策略，减少干扰值影响
3. 【NEW 评论采集】新增 fetch_reviews() 方法，支持：
     - 仅采差评（1-2星，用于竞品分析）
     - 全部评论
     - 每个商品可采集 20~30 条（2-3页，每页10条）
     输出到 Excel 新增"客户差评"列（前3条摘要 + 总条数）

架构：搜索页只取基础字段 → 详情页采集所有深度字段（含价格）→ 评论页采集差评
"""

import sys
import os
import re
import time
import random
import urllib.parse
import io
import json
from datetime import datetime

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image


# ════════════════════════════════════════════════════════════
#  配置
# ════════════════════════════════════════════════════════════

BASE_URL = "https://www.amazon.com/s"

HEADERS_POOL = [
    {
        "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/124.0.0.0 Safari/537.36"),
        "Accept-Language": "en-US,en;q=0.9",
    },
    {
        "User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                       "AppleWebKit/605.1.15 (KHTML, like Gecko) "
                       "Version/17.5 Safari/605.1.15"),
        "Accept-Language": "en-US,en;q=0.8",
    },
    {
        "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) "
                       "Gecko/20100101 Firefox/125.0"),
        "Accept-Language": "en-US,en;q=0.7",
    },
]

COMMON_HEADERS = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "DNT": "1",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Accept-Language": "en-US,en;q=0.9",
}

SORT_MAP = {
    "sales":   "exact-aware-popularity-rank",
    "reviews": "review-rank",
}

# ── 评论采集配置 ──
REVIEW_STAR_MAP = {
    "negative":  "critical",      # 1-2星 差评（推荐：用于竞品分析）
    "all":       "",              # 全部评论
    "one_star":  "one_star",      # 仅1星
    "two_star":  "two_star",      # 仅2星
    "three_star":"three_star",    # 仅3星（中性）
    "positive":  "positive",      # 4-5星
}


# ════════════════════════════════════════════════════════════
#  解析函数
# ════════════════════════════════════════════════════════════

def parse_price(text):
    """
    价格解析：提取 $12.99 格式的美元价格
    只匹配以 $ 开头的标准美元格式
    """
    if not text or not isinstance(text, str):
        return None, None
    text = text.strip()
    if not text:
        return None, None

    # 匹配 $12.99 或 $1,299.00 格式
    m = re.search(r'\$[\d,]+\.\d{2}', text)
    if m:
        original = m.group()
        num_text = original.replace('$', '').replace(',', '')
        try:
            val = float(num_text)
            if val >= 0.01:
                return val, original
        except ValueError:
            pass

    # 备用：匹配 $12 格式（无小数部分）
    m = re.search(r'\$(\d{1,6})(?!\d)', text)
    if m:
        original = m.group()
        try:
            val = int(m.group(1))
            if val >= 1:
                return float(val), original
        except ValueError:
            pass

    return None, None


def parse_sales(text):
    """月销量解析：10K+ bought in past month 等"""
    if not text or not isinstance(text, str):
        return None, None
    text = text.strip()
    if not text:
        return None, None

    m = re.search(r'([\d,]+\.?\d*[KMB]?\+?)\s*(?:bought|purchased)\s+in\s+past\s+month', text, re.IGNORECASE)
    if m:
        orig = m.group(1)
        s = orig.replace('+', '').replace(',', '').upper()
        try:
            if s.endswith('K'): v = float(s[:-1]) * 1000
            elif s.endswith('M'): v = float(s[:-1]) * 1000000
            elif s.endswith('B'): v = float(s[:-1]) * 1000000000
            else: v = float(s)
            return int(v), orig
        except ValueError:
            pass
    return None, None


def parse_reviews(text):
    """评论数解析：(4,310) 括号内数字"""
    if not text or not isinstance(text, str):
        return None, None
    text = text.strip()
    m = re.search(r'\(([\d,]+)\)', text)
    if m:
        return int(m.group(1).replace(',', '')), f"({m.group(1)})"
    m = re.search(r'([\d,]+)\s*(?:ratings?|reviews?)', text, re.IGNORECASE)
    if m:
        return int(m.group(1).replace(',', '')), m.group(1)
    return None, None


def smart_delay(base=2.5, extra=0.0):
    delay = base + extra + random.uniform(0.5, 1.5)
    time.sleep(delay)


def download_image(url, save_path, session):
    try:
        resp = session.get(url, timeout=20)
        if resp.status_code != 200:
            return False
        img = Image.open(io.BytesIO(resp.content)).convert("RGB")
        img.save(save_path, "PNG", optimize=True, quality=95)
        return True
    except Exception as e:
        print(f"    [图片] 下载失败: {e}")
        return False


# ════════════════════════════════════════════════════════════
#  采集器类
# ════════════════════════════════════════════════════════════

class AmazonScraper:

    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update(COMMON_HEADERS)
        self._rotate_ua()

    def _rotate_ua(self):
        self.session.headers.update(random.choice(HEADERS_POOL))

    # ── 页面请求 ──────────────────────────────────────────────

    def fetch(self, url, max_retries=5):
        for attempt in range(1, max_retries + 1):
            try:
                if attempt > 1:
                    backoff = min(2 ** (attempt - 2), 8)
                    print(f"    [重试] 第 {attempt} 次（等待 {backoff}s）")
                    time.sleep(backoff)
                    self._rotate_ua()

                resp = self.session.get(url, timeout=15)
                resp.raise_for_status()

                if "captcha" in resp.text.lower() or "type the characters" in resp.text.lower():
                    print(f"    [警告] 触发 CAPTCHA...")
                    time.sleep(10 + random.uniform(5, 10))
                    continue
                return resp

            except requests.exceptions.HTTPError as e:
                if e.response.status_code == 429:
                    print(f"    [警告] 限流 429...")
                    time.sleep(15 + random.uniform(5, 10))
                    continue
                elif e.response.status_code == 503:
                    time.sleep(10 + random.uniform(3, 7))
                    continue
                elif attempt == max_retries:
                    raise
            except (requests.exceptions.RequestException, Exception) as e:
                print(f"    [错误] 请求失败: {e}")
                if attempt == max_retries:
                    raise
        return None

    # ── 搜索页面解析（仅基础字段）────────────────────────────

    def search_page(self, keyword, sort_by, page=1):
        print(f"  [搜索] 正在采集第 {page} 页...")

        params = {
            "k": keyword,
            "s": SORT_MAP.get(sort_by, SORT_MAP["sales"]),
            "page": page,
            "language": "en_US",
            "currency": "USD",
        }
        url = f"{BASE_URL}?{urllib.parse.urlencode(params)}"
        resp = self.fetch(url)
        if not resp:
            return []

        soup = BeautifulSoup(resp.text, "html.parser")
        items = soup.select("div[data-component-type='s-search-result']")
        if not items:
            items = soup.select("div[data-asin]")

        print(f"    [结果] 找到 {len(items)} 个商品项")
        products = []
        for idx, item in enumerate(items[:16]):
            product = self._parse_search_item(item)
            if product:
                products.append(product)
        return products

    def _parse_search_item(self, item):
        """
        搜索结果页仅提取：标题、URL、图片、评论数、月销量
        价格和其余全部字段由详情页采集（更准确）
        """
        try:
            # 标题
            title = None
            h2_tag = item.select_one("h2 a span")
            if h2_tag:
                title = h2_tag.get_text(strip=True)
            else:
                title_tag = item.select_one("h2")
                if title_tag:
                    title = title_tag.get_text(strip=True)

            # URL
            url = None
            for selector in ["h2 a", "a.a-link-normal", "a[href*='/dp/']"]:
                a_tag = item.select_one(selector)
                if a_tag and a_tag.has_attr("href"):
                    href = a_tag["href"]
                    if href and href.strip():
                        url = ("https://www.amazon.com" + href) if href.startswith("/") else href
                        break

            # 评论数
            reviews = None
            rating_link = item.select_one("a[href*='#customerReviews'], a[href*='reviewsCount']")
            if rating_link:
                aria = rating_link.get("aria-label") or ""
                link_txt = rating_link.get_text(strip=True) or ""
                r, rt = parse_reviews(aria or link_txt)
                if r:
                    reviews = r
            if reviews is None:
                for span in item.select("span"):
                    txt = span.get_text(strip=True)
                    m = re.search(r'\(([\d,]+)\)', txt)
                    if m:
                        reviews = int(m.group(1).replace(',', ''))
                        break

            # 月销量
            monthly_sales = None
            for elem in item.select("span, div, a"):
                txt = elem.get_text(strip=True).lower()
                if "bought" in txt and "month" in txt:
                    raw = elem.get_text(strip=True)
                    s, st = parse_sales(raw)
                    if s:
                        monthly_sales = s
                        break

            # 图片 URL
            img_url = None
            img_tag = item.select_one("img.s-image") or item.select_one("img")
            if img_tag:
                if img_tag.has_attr("src"):
                    img_url = img_tag["src"]
                elif img_tag.has_attr("data-src"):
                    img_url = img_tag["data-src"]
            if img_url:
                img_url = re.sub(r'\._[A-Z0-9_,]+_', '._SL500_', img_url)
                img_url = re.sub(r'\._AC_', '._AC_SL500_', img_url)

            if not title:
                return None

            return {
                "title":         title,
                "url":           url,
                "reviews":       reviews,
                "monthly_sales": monthly_sales,
                "img_url":       img_url,
                # 以下字段由详情页填充
                "price":         None,
                "price_text":    None,
                "asin":          None,
                "rating":        None,
                "bsr":           None,
                "bullets":       None,
                "specs":         None,
                "brand_name":    None,
                "img_local":     None,
                # 评论字段（可选）
                "negative_reviews": None,
            }

        except Exception as e:
            print(f"  [解析] 单条商品出错: {e}")
            return None

    # ══════════════════════════════════════════════════════
    #  详情页全字段采集（v6.3 核心）
    # ══════════════════════════════════════════════════════

    def fetch_detail_page_full(self, url):
        """
        访问商品详情页，采集所有字段
        """
        try:
            if "?" in url:
                detail_url = url + "&language=en_US&currency=USD"
            else:
                detail_url = url + "?language=en_US&currency=USD"

            resp = self.fetch(detail_url)
            if not resp:
                return {}

            soup = BeautifulSoup(resp.text, "html.parser")
            html_text = str(soup)
            result = {}

            result["price"], result["price_text"] = self._extract_price_v63(soup, html_text)
            result["asin"] = self._extract_asin(soup, url)
            result["rating"] = self._extract_rating_v61(soup)
            result["bsr"] = self._extract_bsr_merged(soup)
            result["bullets"] = self._extract_bullets(soup)
            result["specs"] = self._extract_specs_enhanced(soup)
            result["brand_name"] = self._extract_brand_name_v61(soup)

            return result

        except Exception as e:
            print(f"    [详情页] 全字段采集失败: {e}")
            return {}

    # ── [v6.3 核心修复] 价格提取 — 精准定位 BuyBox 主价格 ──

    def _extract_price_v63(self, soup, html_text=None):
        """
        v6.3 价格提取（修复版）

        核心问题：v6.2 的 Tier 0b 中 `.a-price .a-offscreen` 命中面过广，
        Amazon 页面有多个 .a-price（划线原价、分期价、配件价），select_one
        只取第一个，不一定是 BuyBox 主价格。

        v6.3 修复：
        - 新增 Tier 0 Premium：专门针对 corePriceDisplay、priceToPay 等
          Amazon 最新版本的核心价格容器，精准性最高
        - 调整 .a-price .a-offscreen 的使用方式：
          必须在 BuyBox 专属容器内才命中，不允许全局扫描
        - Tier 3 兜底改为"中位数策略"，避免配件低价干扰

        选择器优先级（从高到低）：
          Tier 0 Premium → Tier 0a → Tier 0b（限制范围版）→
          Tier 0c → Tier 1 → Tier 2 → Tier 3
        """
        if html_text is None:
            html_text = str(soup)

        # ═══════ Tier 0 Premium：Amazon 最新版核心价格容器（最精准）═══════
        # 这些是 Amazon 当前（2024-2026）产品页的主流价格 ID，
        # 专属 BuyBox 区域，不会包含划线价、配件价
        premium_selectors = [
            # 最新版：corePriceDisplay_desktop 专用容器
            "#corePriceDisplay_desktop_feature_div .a-offscreen",
            "#corePriceDisplay_desktop_feature_div span.a-price-whole",
            # priceToPay — BuyBox 实付价格（最可靠之一）
            ".priceToPay span.a-offscreen",
            ".priceToPay .a-price-whole",
            # reinventPricePolicyMessage 下的核心价格
            "#reinventPricePolicyMessage .a-offscreen",
            # apex_offerDisplay — 另一种新版价格容器
            "#apex_offerDisplay .a-offscreen",
        ]
        for selector in premium_selectors:
            el = soup.select_one(selector)
            if el:
                raw = el.get_text(strip=True).strip().replace("\xa0", " ")
                # 若是 whole 部分，需拼接小数
                if "whole" in selector:
                    parent = el.find_parent(class_=re.compile(r"a-price"))
                    if parent:
                        frac = parent.select_one(".a-price-fraction")
                        whole_txt = re.sub(r'[^\d]', '', el.get_text(strip=True))
                        frac_txt = re.sub(r'[^\d]', '', frac.get_text(strip=True)) if frac else "00"
                        raw = f"${whole_txt}.{frac_txt}"
                val, fmt = self._parse_usd_price(raw)
                if val is not None:
                    print(f"    [价格-TierP-{selector[:50]}] {fmt}")
                    return val, fmt

        # ═══════ Tier 0a: corePrice_feature_div 容器（次精准）═══════
        for selector in [
            "#corePrice_feature_div .a-offscreen",
            "#corePrice_desktop .a-offscreen",
            "#corePrice .a-offscreen",
            "div#corePrice_feature_div span.a-offscreen",
        ]:
            el = soup.select_one(selector)
            if el:
                raw = el.get_text(strip=True).strip().replace("\xa0", " ")
                val, fmt = self._parse_usd_price(raw)
                if val is not None:
                    print(f"    [价格-Tier0a] {fmt}")
                    return val, fmt

        # ═══════ Tier 0b: priceblock 经典 ID（旧版页面，不会有歧义）═══════
        for pid in ["#priceblock_ourprice", "#priceblock_dealprice"]:
            el = soup.select_one(pid)
            if el:
                raw = el.get_text(strip=True).strip().replace("\xa0", " ")
                val, fmt = self._parse_usd_price(raw)
                if val is not None:
                    print(f"    [价格-Tier0b-{pid}] {fmt}")
                    return val, fmt

        # ═══════ Tier 0b+: .a-price .a-offscreen —— 严格限制在 BuyBox 内 ═══════
        # 不允许全局命中，必须在已知的 BuyBox 容器内
        buybox_containers = [
            "#buyBoxSection",
            "#buybox",
            "#desktop_buybox",
            "#desktop_unifiedPrice",
            "#unifiedPrice_feature_div",
        ]
        for bb_sel in buybox_containers:
            bb = soup.select_one(bb_sel)
            if bb:
                # 在 BuyBox 内找 .a-offscreen，但跳过 .a-text-strike（划线原价）
                for off_el in bb.select("span.a-offscreen"):
                    # 检查父元素是否是划线价
                    parent = off_el.find_parent()
                    if parent and "strike" in str(parent.get("class", "")):
                        continue
                    grandparent = off_el.find_parent(class_=re.compile(r"a-text-strike|a-color-secondary"))
                    if grandparent:
                        continue
                    raw = off_el.get_text(strip=True).strip().replace("\xa0", " ")
                    val, fmt = self._parse_usd_price(raw)
                    if val is not None:
                        print(f"    [价格-Tier0b+({bb_sel})] {fmt}")
                        return val, fmt

        # ═══════ Tier 0c: BuyBox 整体文本正则（二次兜底）═══════
        for bb_sel in buybox_containers:
            bb = soup.select_one(bb_sel)
            if bb:
                bb_text = bb.get_text()
                m = re.search(r'\$([\d,]+\.\d{2})', bb_text)
                if m:
                    num = m.group(1).replace(",", "")
                    try:
                        val = float(num)
                        if 0.50 <= val <= 50000:
                            fmt = f"${m.group(1)}"
                            print(f"    [价格-Tier0c-buyBox-text] {fmt}")
                            return val, fmt
                    except ValueError:
                        pass

        # ═══════ Tier 1: 价格相关 ID/class 元素 ═══════
        price_selectors = [
            "#apex_desktop_price", "#apex_price", ".apexPrice",
            ".price-inside-feature-div",
            "[id*='Price'] .a-offscreen",
        ]
        for sel in price_selectors:
            el = soup.select_one(sel)
            if el:
                raw = el.get_text(strip=True).strip().replace("\xa0", " ")
                val, fmt = self._parse_usd_price(raw)
                if val is not None:
                    print(f"    [价格-Tier1] {fmt}")
                    return val, fmt

        # Tier 1b: a-price-whole + a-price-fraction 拼接（必须排除划线价）
        for price_el in soup.select(".a-price"):
            # 跳过划线原价容器
            classes = " ".join(price_el.get("class", []))
            if "strike" in classes or "secondary" in classes:
                continue
            whole = price_el.select_one(".a-price-whole")
            frac = price_el.select_one(".a-price-fraction")
            if whole and frac:
                whole_txt = re.sub(r'[^\d]', '', whole.get_text(strip=True))
                frac_txt = re.sub(r'[^\d]', '', frac.get_text(strip=True))
                if whole_txt and frac_txt:
                    raw = f"${whole_txt}.{frac_txt}"
                    val, fmt = self._parse_usd_price(raw)
                    if val is not None:
                        print(f"    [价格-Tier1b-拼接] {fmt}")
                        return val, fmt

        # ═══════ Tier 2: LD+JSON structured data ═══════
        for s in soup.find_all("script", type="application/ld+json"):
            try:
                d = json.loads(s.string)
                offers = d.get("offers", []) if isinstance(d, dict) else []
                if isinstance(offers, dict):
                    offers = [offers]
                for o in offers:
                    p = o.get("price")
                    cur = o.get("priceCurrency", "").upper()
                    if p and ("USD" in cur or "US" in cur or cur == "" or cur is None):
                        try:
                            v = float(p)
                            if 0.50 <= v <= 50000:
                                fmt = f"${p}"
                                print(f"    [价格-Tier2-LDJSON] {fmt}")
                                return v, fmt
                        except (ValueError, TypeError):
                            pass
            except (json.JSONDecodeError, AttributeError):
                pass

        # ═══════ Tier 3: 全文正则兜底（中位数策略，减少异常值干扰）═══════
        candidates = []
        for m in re.finditer(r'\$([\d,]+\.\d{2})', html_text):
            num = m.group(1).replace(",", "")
            try:
                val = float(num)
                # 过滤明显不合理的价格（< $1 或 > $5000 极端值）
                if 1.0 <= val <= 5000:
                    fmt = f"${m.group(1)}"
                    candidates.append((val, fmt))
            except ValueError:
                pass

        if candidates:
            # 去重
            seen = set()
            unique = []
            for v, f in candidates:
                if f not in seen:
                    seen.add(f)
                    unique.append((v, f))

            if len(unique) == 1:
                print(f"    [价格-Tier3-唯一] {unique[0][1]}")
                return unique[0]

            # 中位数策略：排序后取中间值，比取最低值更不易受干扰
            unique.sort(key=lambda x: x[0])
            mid_idx = len(unique) // 2
            # 优先取 $10-$500 区间（最常见的消费品价格区间）
            common_range = [(v, f) for v, f in unique if 5.0 <= v <= 500.0]
            if common_range:
                # 在合理区间内取中位数
                mid_idx = len(common_range) // 2
                best_val, best_fmt = common_range[mid_idx]
            else:
                best_val, best_fmt = unique[mid_idx]

            print(f"    [价格-Tier3-中位数] {best_fmt} (从{len(unique)}个唯一候选中选)")
            return best_val, best_fmt

        print(f"    [价格] 未找到有效美区USD价格")
        return None, None

    def _parse_usd_price(self, text):
        """
        解析 USD 价格文本，返回 (float_value, formatted_string) 或 (None, None)
        只接受标准 $XX.XX 格式
        """
        if not text or not isinstance(text, str):
            return None, None
        text = text.strip().replace("\xa0", " ")

        # 标准 $12.99 或 $1,299.00 格式
        m = re.match(r'^\$([\d,]+\.\d{2})$', text)
        if m:
            num_str = m.group(1).replace(",", "")
            try:
                val = float(num_str)
                if 0.01 <= val <= 99999.99:
                    return val, f"${m.group(1)}"
            except ValueError:
                pass

        # 无小数的 $12 格式
        m = re.match(r'^\$(\d{1,6})$', text)
        if m:
            try:
                val = float(m.group(1))
                if 1 <= val <= 99999:
                    return val, text
            except ValueError:
                pass

        return None, None

    # ── [Fix 2] 评分提取 v6.1 ─────────────────────────────

    def _extract_rating_v61(self, soup):
        """
        提取星级评分数字（如 4.8）
        """
        # 方式1: 直接找 avgCustomerReviewsFeatureDecimal 附近（最精准）
        decimal_elem = soup.select_one(
            "#averageCustomerReviewsFeatureDecimal, "
            "[id*='avgRating'], "
            ".a-size-base.a-color-base"
        )
        if decimal_elem:
            txt = decimal_elem.get_text(strip=True)
            m = re.match(r'^([\d.]+)$', txt)
            if m:
                val = float(m.group(1))
                if 1.0 <= val <= 5.0:
                    print(f"    [评分-decimal] {val}")
                    return val

        # 方式2: 找包含 "out of 5 stars" 的 aria-label
        for el in soup.select("[aria-label], [title]"):
            label = (el.get("aria-label") or el.get("title") or "")
            m = re.search(r'([\d.]+)\s*out\s+of\s+5\s+stars', label, re.IGNORECASE)
            if m:
                val = float(m.group(1))
                if 1.0 <= val <= 5.0:
                    print(f"    [评分-aria] {val}")
                    return val

        # 方式3: 在页面文本中找 "4.8 ★★★★☆ (415)" 模式
        body = soup.body.get_text() if soup.body else ""
        m = re.search(r'(^|\n)\s*(\d\.\d)\s*(?:★|⭐|out\s+of\s*5)', body)
        if m:
            val = float(m.group(2))
            if 1.0 <= val <= 5.0:
                print(f"    [评分-text] {val}")
                return val

        # 方式4: 从 a-star-* CSS类反推
        for el in soup.select("[class*='a-star-']"):
            for cls in el.get("class", []):
                if cls.startswith("a-star-") and cls.count("-") >= 2:
                    parts = cls.replace("a-star-", "").split("-")
                    if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
                        try:
                            val = float(f"{parts[0]}.{parts[1]}")
                            if 1.0 <= val <= 5.0:
                                print(f"    [评分-class] {val}")
                                return val
                        except ValueError:
                            pass

        return None

    # ── ASIN 提取 ──────────────────────────────────────────

    def _extract_asin(self, soup, url):
        # 方式1: input[name='ASIN']
        inp = soup.select_one("#ASIN input[name='ASIN']")
        if inp and inp.get("value"):
            v = inp["value"].strip()
            if re.match(r'^[A-Z0-9]{10}$', v):
                return v

        # 方式2: data-asin 属性
        da = soup.select_one("[data-asin]")
        if da:
            v = (da.get("data-asin") or "").strip()
            if re.match(r'^[A-Z0-9]{10}$', v):
                return v

        # 方式3: URL 中提取
        m = re.search(r'/dp/([A-Z0-9]{10})', url or "")
        if m:
            return m.group(1)

        # 方式4: 文本中查找
        body = soup.get_text() if soup else ""
        m = re.search(r'ASIN[:\s]+([A-Z0-9]{10})', body)
        if m:
            return m.group(1)

        return None

    # ── BSR 全量排名提取 ────────────────────────────────────

    def _extract_bsr_merged(self, soup):
        """
        从 "Best Sellers Rank" 区域提取所有大类 + 小类完整排名信息
        输出格式: "#660 in Home & Kitchen | #1 in Collectible Dolls"
        """
        ranks = []

        bsr_containers = [
            "#productDetails_detailBullets_sections1",
            "#productDetails_db_sections",
            "table#productDetails",
            "table.productDetailsDetail",
            "#SalesRank",
            "#detailBulletsWrapper_feature_div",
            "#detailBullets_feature_div",
            ".product-details-table",
        ]

        for sel in bsr_containers:
            container = soup.select_one(sel)
            if container:
                rows = container.select("tr, li")
                for row in rows:
                    row_text = row.get_text(strip=True)
                    if "#" not in row_text:
                        continue
                    matches = re.findall(r'#([\d,]+)\s+(?:in|en)\s+([^\n\r\(]+?)(?:\s*\(|$)', row_text, re.IGNORECASE)
                    for rank_num, cat in matches:
                        cat_clean = cat.strip().rstrip("(").strip()
                        entry = f"#{rank_num} in {cat_clean}"
                        if entry not in ranks and len(cat_clean) > 1 and len(cat_clean) < 100:
                            ranks.append(entry)

        # 全文正则补充
        body = soup.body.get_text() if soup.body else ""
        matches = re.findall(r'(#[\d,]+)\s+in\s+([^\n\r]+?)(?:\s*\(|\s*$|\s*(?:See|in))', body, re.IGNORECASE)
        seen_cats = set()
        for rank_num, cat in matches:
            cat_clean = cat.split("(")[0].strip().rstrip(",").strip()
            if cat_clean and cat_clean not in seen_cats and len(cat_clean) < 100:
                skip_words = ["best sellers", "rank", "top", "rated", "customer reviews"]
                is_valid = not any(sw in cat_clean.lower() for sw in skip_words)
                if is_valid:
                    seen_cats.add(cat_clean)
                    entry = f"#{rank_num} in {cat_clean}"
                    if entry not in ranks:
                        ranks.append(entry)

        if ranks:
            result = " | ".join(ranks[:8])
            print(f"    [BSR] 共{len(ranks)}条: {result[:100]}...")
            return result

        return None

    # ── 五点描述提取 ────────────────────────────────────────

    def _extract_bullets(self, soup):
        bullets = []

        ul = soup.select_one("#feature-bullets ul, #feature-bullets, "
                             ".a-unordered-list.a-vertical.a-spacing-mini")
        if ul:
            for li in ul.select("li"):
                txt = li.get_text(strip=True)
                if txt and len(txt) > 8:
                    bullets.append(txt)

        if not bullets:
            sec = soup.select_one("#about-this-item-content, #productDescription")
            if sec:
                for tag in sec.select("p, li, div"):
                    txt = tag.get_text(strip=True)
                    if txt and len(txt) > 12:
                        bullets.append(txt)

        if not bullets:
            ap = soup.select_one("#aplus, #aplus_feature_div, "
                               "div[data-feature-name='aplus-standard']")
            if ap:
                for tag in ap.select("p, li"):
                    txt = tag.get_text(strip=True)
                    if txt and len(txt) > 12:
                        bullets.append(txt)

        return "\n".join(bullets[:5]) if bullets else None

    # ── 规格参数精简提取 ─────────────────────────────────────

    def _extract_specs_enhanced(self, soup):
        """
        精简增强版：仅提取三类关键规格参数
          1. 重量 (Weight)
          2. 尺寸 (Dimension)
          3. 材质 (Material)
        """
        specs = {}

        prod_details = soup.select_one("#prodDetails")
        if prod_details:
            self._parse_proddetails_container(prod_details, specs)

        for container in soup.select(
            "#productDetails_detailBullets_sections1, "
            "#productDetails_db_sections, "
            "table#productDetails, table.productDetailsDetail, "
            "#techDetails_techSpec_section_1"
        ):
            self._parse_table_rows(container, specs)

        for sel in ["#technicalDetails_list", "#detailBulletsWrapper_feature_div"]:
            el = soup.select_one(sel)
            if el:
                self._parse_kv_elements(el, specs)

        aplus_tables = soup.select("#aplus table, #aplus_feature_div table, "
                                   "div.aplus-module table, [data-feature-name*='aplus'] table")
        for tbl in aplus_tables:
            self._parse_table_rows(tbl, specs)

        bullets_ul = soup.select_one("#feature-bullets ul, #feature-bullets, "
                                     ".a-unordered-list.a-vertical.a-spacing-mini")
        if bullets_ul:
            for li in bullets_ul.select("li"):
                txt = li.get_text(strip=True)
                dim_match = re.search(
                    r'(\d+\.?\d*)\s*(?:inches?|inch|in|cm|mm)\s*(?:tall|x|by|wide|long|high|deep|diameter)',
                    txt, re.IGNORECASE)
                if dim_match:
                    ctx_start = max(0, txt.lower().find(dim_match.group(0).lower()) - 30)
                    spec_txt = txt[ctx_start:ctx_start + len(dim_match.group(0)) + 40].strip()
                    if "bullet_dimension" not in specs:
                        specs["bullet_dimension"] = spec_txt

                mat_match = re.search(
                    r'(?:made\s+(?:of|from)|craft(?:ed)?\s+from|material|fabric)\s*[:is]+?\s*([A-Za-z][A-Za-z\s\-\/,()]{2,60}?)(?:\.|,|$|and|with|for)',
                    txt, re.IGNORECASE)
                if mat_match and "material" not in specs:
                    specs["bullet_material"] = mat_match.group(1).strip()

        title = soup.find("meta", attrs={"property": "og:title"})
        title_text = (title.get("content", "") if title else "")
        if not title_text:
            t = soup.select_one("#productTitle")
            if t:
                title_text = t.get_text(strip=True)

        size_m = re.search(r'(\d+\.?\d*)"\s*[x×]\s*(\d+\.?\d*)"(?:\s*[x×]\s*(\d+\.?\d*)")?', title_text)
        if size_m and "dimension" not in specs:
            specs["title_dimension"] = size_m.group(0)

        if specs:
            lines = []
            ordered_keys = []
            for cat_keywords in [["weight"], ["dimension", "size"], ["material"]]:
                for sk in list(specs.keys()):
                    if any(kw in sk for kw in cat_keywords) and sk not in ordered_keys:
                        ordered_keys.append(sk)

            for sk in ordered_keys:
                lines.append(f"{sk}: {specs[sk]}")

            result = "\n".join(lines[:10])
            if result:
                print(f"    [规格] {len(lines)} 条 (W/D/M 精简)")
            return result

        return None

    def _parse_proddetails_container(self, container, specs):
        rows = container.select("tr")
        if rows:
            for row in rows:
                cells = row.select("th, td")
                if len(cells) >= 2:
                    key = cells[0].get_text(strip=True).rstrip(": ").lower()
                    val = cells[1].get_text(strip=True)
                    if key and val and self._is_target_spec(key) and key not in specs:
                        specs[key] = val
                elif len(cells) == 1:
                    txt = cells[0].get_text(strip=True)
                    if ":" in txt:
                        k, v = txt.split(":", 1)
                        k_clean = k.strip().lower()
                        if self._is_target_spec(k_clean) and k_clean not in specs:
                            specs[k_clean] = v.strip()

        container_text = container.get_text(separator='|', strip=True)

        kv_patterns = [
            r'(item\s+dimensions?\s*l?\s*x?\s*w?\s*x?\s*h?)\s{2,}([^\n\r|]+)',
            r'(package\s+dimensions?)\s{2,}([^\n\r|]+)',
            r'(item\s+weight)\s{2,}([^\n\r|]+)',
            r'(package\s+weight)\s{2,}([^\n\r|]+)',
            r'(material\s+type)\s{2,}([^\n\r|]+)',
            r'(outer\s+material)\s{2,}([^\n\r|]+)',
            r'\b(material)\b\s{2,}([^\n\r|]{2,60})',
            r'(item\s+dimensions?.*?):\s*([^\n\r|]+)',
            r'(package\s+dimensions?.*?):\s*([^\n\r|]+)',
            r'(item\s+weight):\s*([^\n\r|]+)',
            r'(package\s+weight):\s*([^\n\r|]+)',
            r'(material\s+type):\s*([^\n\r|]+)',
            r'(outer\s+material):\s*([^\n\r|]+)',
        ]
        for pattern in kv_patterns:
            m = re.search(pattern, container_text, re.IGNORECASE)
            if m:
                k = m.group(1).strip().lower()
                v = m.group(2).strip()
                if v and self._is_target_spec(k) and k not in specs:
                    v = re.sub(r'\s+', ' ', v).strip()
                    v = re.sub(r'^[\|\s]+', '', v).strip()
                    if v and len(v) > 1 and len(v) < 200:
                        specs[k] = v

        for child in container.children:
            if hasattr(child, 'get_text'):
                child_text = child.get_text(strip=True)
                lower = child_text.lower()
                has_target = any(kw in lower for kw in [
                    'dimension', 'weight', 'material type', 'outer material'
                ])
                if has_target and ':' in child_text:
                    parts = child_text.split(':', 1)
                    if len(parts) == 2:
                        k = parts[0].strip().lower()
                        v = parts[1].strip()
                        if self._is_target_spec(k) and k not in specs and v:
                            specs[k] = v

    def _parse_table_rows(self, container, specs):
        for row in container.select("tr"):
            th = row.select_one("th")
            td = row.select_one("td")
            if th and td:
                key = th.get_text(strip=True).rstrip(": ").lower()
                val = td.get_text(strip=True)
                if key and val and self._is_target_spec(key) and key not in specs:
                    specs[key] = val

    def _parse_kv_elements(self, container, specs):
        for item in container.select("tr, li, dt"):
            txt = item.get_text(strip=True)
            if ":" in txt:
                k, v = txt.split(":", 1)
                k_clean = k.strip().lower()
                if self._is_target_spec(k_clean) and k_clean not in specs:
                    specs[k_clean] = v.strip()

    def _is_target_spec(self, key_lower):
        target_words = ["weight", "dimension", "size", "material"]
        return any(tw in key_lower for tw in target_words)

    # ── 品牌提取 ──────────────────────────────────────────────

    def _extract_brand_name_v61(self, soup):
        # 方式1: bylineInfo
        brand_link = soup.select_one("#bylineInfo, a#bylineInfo")
        if brand_link:
            txt = brand_link.get_text(strip=True)
            txt = re.sub(r'^\s*(?:Brand|品牌|Visit\s+the\s+|Store)[:\s]*', '', txt, flags=re.IGNORECASE)
            if txt and len(txt) < 200:
                print(f"    [品牌-bylineInfo] {txt}")
                return txt

        # 方式2: Visit the xxx Store
        store_links = soup.select("a[href*='/store/'], a[href*='/merchant-items/']")
        for sl in store_links:
            txt = sl.get_text(strip=True)
            m = re.search(r'Visit\s+the\s+([A-Za-z0-9\s\-\.&\']+?)\s*Store', txt, re.IGNORECASE)
            if m:
                brand = m.group(1).strip()
                if brand and len(brand) < 100:
                    print(f"    [品牌-StoreLink] {brand}")
                    return brand

        # 方式3: meta 标签
        for meta in soup.find_all("meta"):
            prop = (meta.get("property") or meta.get("name") or "").lower()
            if prop in ("og:brand", "brand"):
                content = meta.get("content", "").strip()
                if content:
                    print(f"    [品牌-meta] {content}")
                    return content

        # 方式4: Sold by
        body = soup.get_text()
        m = re.search(r'Sold by\s+([A-Za-z0-9\s\-\.&\']+?)(?:\s*and|Fulfilled| Ships|$|\n)', body)
        if m:
            seller = m.group(1).strip()
            if len(seller) > 2 and len(seller) < 200:
                print(f"    [品牌-SoldBy] {seller}")
                return seller

        return None

    # ══════════════════════════════════════════════════════
    #  [NEW v6.3] 评论采集功能
    #  支持：仅差评（1-2星）/ 全部评论
    #  每商品可采集 20~30 条（2~3 页，每页 10 条）
    # ══════════════════════════════════════════════════════

    def fetch_reviews(self, asin, star_filter="negative", max_pages=2):
        """
        采集指定 ASIN 的评论内容

        参数：
          asin        - 10位 Amazon 商品识别码
          star_filter - 筛选类型：
                          "negative"  → 差评（1-2星，最常用）
                          "one_star"  → 仅1星
                          "two_star"  → 仅2星
                          "all"       → 全部评论（不过滤）
                          "positive"  → 好评（4-5星）
          max_pages   - 最多采集页数（每页10条，推荐2-3页）

        返回：
          list of dict: [{
            "star": "1.0 out of 5 stars",
            "title": "评论标题",
            "body": "评论正文（前300字）",
            "date": "Reviewed in the United States on ...",
            "helpful": "X people found this helpful",
            "verified": True/False,
          }]
        """
        if not asin:
            return []

        filter_param = REVIEW_STAR_MAP.get(star_filter, "critical")
        reviews = []

        for page in range(1, max_pages + 1):
            try:
                # Amazon 评论页 URL 结构
                params = {
                    "pageNumber": page,
                    "sortBy": "recent",          # 按最近排序，差评更新鲜
                }
                if filter_param:
                    params["filterByStar"] = filter_param

                url = (f"https://www.amazon.com/product-reviews/{asin}/?"
                       + urllib.parse.urlencode(params))

                print(f"    [评论-{star_filter}] 第{page}页: {url[:80]}...")
                resp = self.fetch(url)
                if not resp:
                    break

                soup = BeautifulSoup(resp.text, "html.parser")
                page_reviews = self._parse_review_page(soup)

                if not page_reviews:
                    print(f"    [评论] 第{page}页无数据，停止")
                    break

                reviews.extend(page_reviews)
                print(f"    [评论] 第{page}页采集 {len(page_reviews)} 条，累计 {len(reviews)} 条")

                # 检查是否还有下一页
                next_btn = soup.select_one("li.a-last a, .a-pagination .a-last:not(.a-disabled) a")
                if not next_btn and page < max_pages:
                    print(f"    [评论] 已到最后一页")
                    break

                if page < max_pages:
                    smart_delay(base=2.0)

            except Exception as e:
                print(f"    [评论] 第{page}页采集失败: {e}")
                break

        return reviews

    def _parse_review_page(self, soup):
        """
        解析评论页面，提取评论列表
        """
        reviews = []

        # Amazon 评论容器选择器
        review_containers = soup.select(
            "div[data-hook='review'], "
            ".review, "
            "[id^='customer_review-']"
        )

        for container in review_containers:
            try:
                review = {}

                # 星级
                star_el = container.select_one(
                    "[data-hook='review-star-rating'] .a-icon-alt, "
                    ".review-rating .a-icon-alt, "
                    "[class*='review-rating'] span"
                )
                if star_el:
                    review["star"] = star_el.get_text(strip=True)
                else:
                    # 备用：从 class 推断
                    for el in container.select("[class*='a-star-']"):
                        for cls in el.get("class", []):
                            if cls.startswith("a-star-") and cls != "a-star-mini":
                                review["star"] = cls.replace("a-star-", "").replace("-", ".") + " out of 5"
                                break

                # 评论标题
                title_el = container.select_one(
                    "[data-hook='review-title'] span:not([class*='a-color-secondary']), "
                    ".review-title span"
                )
                if title_el:
                    review["title"] = title_el.get_text(strip=True)

                # 评论正文（截取前 400 字）
                body_el = container.select_one(
                    "[data-hook='review-body'] span, "
                    ".review-text-content span, "
                    ".review-text span"
                )
                if body_el:
                    body_text = body_el.get_text(strip=True)
                    # 清理多余空白
                    body_text = re.sub(r'\s+', ' ', body_text).strip()
                    review["body"] = body_text[:400] + ("..." if len(body_text) > 400 else "")

                # 评论日期
                date_el = container.select_one(
                    "[data-hook='review-date'], "
                    ".review-date"
                )
                if date_el:
                    review["date"] = date_el.get_text(strip=True)

                # "X people found this helpful"
                helpful_el = container.select_one(
                    "[data-hook='helpful-vote-statement'], "
                    ".cr-vote-text"
                )
                if helpful_el:
                    review["helpful"] = helpful_el.get_text(strip=True)

                # 是否认证购买
                verified_el = container.select_one(
                    "[data-hook='avp-badge'], "
                    ".a-color-success"
                )
                review["verified"] = bool(verified_el and "verified" in verified_el.get_text(strip=True).lower())

                # 只保留有正文的评论
                if review.get("body") and len(review["body"]) > 10:
                    reviews.append(review)

            except Exception as e:
                print(f"    [评论解析] 单条出错: {e}")
                continue

        return reviews

    def format_reviews_for_excel(self, reviews, max_show=3):
        """
        将评论列表格式化为 Excel 单元格文本
        格式：前 max_show 条差评摘要 + 总条数

        max_show=3 → 显示前3条，每条：星级 + 标题 + 正文前150字
        """
        if not reviews:
            return None

        lines = [f"共 {len(reviews)} 条差评\n{'─'*30}"]

        for i, r in enumerate(reviews[:max_show], 1):
            parts = []
            if r.get("star"):
                parts.append(f"★ {r['star']}")
            if r.get("title"):
                parts.append(f"【{r['title']}】")
            if r.get("body"):
                body_short = r["body"][:150] + ("..." if len(r["body"]) > 150 else "")
                parts.append(body_short)
            if r.get("date"):
                parts.append(r["date"])
            if parts:
                lines.append(f"\n[差评 {i}]\n" + "\n".join(parts))

        if len(reviews) > max_show:
            lines.append(f"\n（另有 {len(reviews) - max_show} 条差评未显示）")

        return "\n".join(lines)

    # ── 批量采集 ──────────────────────────────────────────────

    def collect(self, keyword, sort_by, total, fetch_reviews_flag=False,
                review_filter="negative", review_pages=2):
        """
        搜索页采集基础字段 → 逐个访问详情页采集全量字段
        可选：采集评论内容

        参数：
          fetch_reviews_flag - 是否采集评论（True/False）
          review_filter      - 评论筛选类型（默认 "negative" 差评）
          review_pages       - 每个商品采集几页评论（每页10条，默认2页=20条）
        """
        all_products = []
        page = 1

        while len(all_products) < total:
            items = self.search_page(keyword, sort_by, page)
            if not items:
                print(f"  [采集] 第 {page} 页无数据，停止")
                break
            all_products.extend(items)
            print(f"  [进度] 已采集 {min(len(all_products), total)} / {total}")
            if len(all_products) >= total:
                break
            page += 1
            smart_delay(base=3.0, extra=(2.0 if page % 5 == 0 else 0.0))

        products = all_products[:total]

        # 详情页全字段采集
        print(f"\n{'='*60}")
        print(f"  [详情页] 开始采集 {len(products)} 个商品的详细数据...")
        print(f"{'='*60}")

        for idx, product in enumerate(products):
            url = product.get("url")
            if url:
                print(f"  [{idx+1}/{len(products)}] {product.get('title', '')[:50]}...")
                detail = self.fetch_detail_page_full(url)
                if detail:
                    for k, v in detail.items():
                        if k in product and v is not None:
                            product[k] = v
                    info_parts = []
                    if detail.get("asin"): info_parts.append(f"ASIN={detail['asin']}")
                    if detail.get("price_text"): info_parts.append(detail['price_text'])
                    if detail.get("rating"): info_parts.append(f"评分={detail['rating']}")
                    if detail.get("brand_name"): info_parts.append(f"品牌={detail['brand_name']}")
                    if detail.get("bsr"): info_parts.append(f"BSR={detail['bsr'][:40]}")
                    if info_parts:
                        print(f"    {' | '.join(info_parts)}")

                # 评论采集（可选）
                if fetch_reviews_flag and product.get("asin"):
                    smart_delay(base=1.5)
                    print(f"    [评论] 采集 ASIN={product['asin']} 的{review_filter}评论...")
                    reviews = self.fetch_reviews(
                        asin=product["asin"],
                        star_filter=review_filter,
                        max_pages=review_pages
                    )
                    if reviews:
                        product["negative_reviews"] = self.format_reviews_for_excel(reviews)
                        print(f"    [评论] 采集到 {len(reviews)} 条，已格式化")
                    else:
                        print(f"    [评论] 未采集到评论数据")

                if idx < len(products) - 1:
                    smart_delay(base=2.0)

        return products


# ════════════════════════════════════════════════════════════
#  Excel 导出 (v6.3 — 新增"客户差评"列)
# ════════════════════════════════════════════════════════════

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(name="Arial", size=11, color="FFFFFF", bold=True)
DATA_FONT    = Font(name="Arial", size=10)
LINK_FONT    = Font(name="Arial", size=9, color="0563C1", underline="single")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN   = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# v6.3 表头（新增"客户差评"列）
COL_SETTINGS_NO_REVIEW = [
    ("序号",      6,   CENTER_ALIGN),
    ("首图",      16,  CENTER_ALIGN),
    ("ASIN",      14,  CENTER_ALIGN),
    ("商品标题",   55,  LEFT_ALIGN),
    ("品牌名称",   18,  CENTER_ALIGN),
    ("价格($)",   11,  CENTER_ALIGN),
    ("评分",       8,   CENTER_ALIGN),
    ("评论数",    10,  CENTER_ALIGN),
    ("BSR排名",    35,  LEFT_ALIGN),
    ("五点描述",   60,  LEFT_ALIGN),
    ("规格参数",   45,  LEFT_ALIGN),
    ("月销量",    12,  CENTER_ALIGN),
    ("商品链接",   50,  LEFT_ALIGN),
]

COL_SETTINGS_WITH_REVIEW = [
    ("序号",      6,   CENTER_ALIGN),
    ("首图",      16,  CENTER_ALIGN),
    ("ASIN",      14,  CENTER_ALIGN),
    ("商品标题",   55,  LEFT_ALIGN),
    ("品牌名称",   18,  CENTER_ALIGN),
    ("价格($)",   11,  CENTER_ALIGN),
    ("评分",       8,   CENTER_ALIGN),
    ("评论数",    10,  CENTER_ALIGN),
    ("BSR排名",    35,  LEFT_ALIGN),
    ("五点描述",   60,  LEFT_ALIGN),
    ("规格参数",   45,  LEFT_ALIGN),
    ("月销量",    12,  CENTER_ALIGN),
    ("客户差评",   80,  LEFT_ALIGN),   # ← 新增列
    ("商品链接",   50,  LEFT_ALIGN),
]


def export_excel(products, output_dir, keyword, with_reviews=False):
    """导出数据到 Excel 文件"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(output_dir, f"amazon_{timestamp}.xlsx")
    images_dir = os.path.join(output_dir, "images")
    os.makedirs(images_dir, exist_ok=True)

    col_settings = COL_SETTINGS_WITH_REVIEW if with_reviews else COL_SETTINGS_NO_REVIEW

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "商品数据"

    # ── 表头 ───────────────────────────────────────────────
    for col_idx, (col_name, col_width, _) in enumerate(col_settings, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # ── 数据行 ─────────────────────────────────────────────
    for row_idx, product in enumerate(products, 2):
        row_height = 120 if with_reviews else 90
        ws.row_dimensions[row_idx].height = row_height
        col = 1

        # 序号
        cell = ws.cell(row=row_idx, column=col, value=row_idx - 1)
        cell.font, cell.alignment, cell.border = DATA_FONT, CENTER_ALIGN, THIN_BORDER
        col += 1

        # 首图
        if product.get("img_url"):
            img_filename = f"{row_idx - 1:03d}.png"
            img_path = os.path.join(images_dir, img_filename)
            sess = requests.Session()
            sess.headers.update(COMMON_HEADERS)
            if download_image(product["img_url"], img_path, sess):
                product["img_local"] = img_path
                try:
                    img = XLImage(img_path)
                    img.width = 85
                    img.height = 85
                    img.anchor = f"B{row_idx}"
                    ws.add_image(img)
                except Exception as e:
                    print(f"    [Excel] 图片插入失败: {e}")
        cell = ws.cell(row=row_idx, column=col, value="")
        cell.alignment, cell.border = CENTER_ALIGN, THIN_BORDER
        col += 1

        # ASIN
        v = product.get("asin")
        cell = ws.cell(row=row_idx, column=col, value=v if v else "null")
        cell.font, cell.alignment, cell.border = DATA_FONT, CENTER_ALIGN, THIN_BORDER
        col += 1

        # 商品标题
        cell = ws.cell(row=row_idx, column=col, value=product.get("title", ""))
        cell.font, cell.alignment, cell.border = DATA_FONT, LEFT_ALIGN, THIN_BORDER
        col += 1

        # 品牌名称
        v = product.get("brand_name")
        cell = ws.cell(row=row_idx, column=col, value=v if v else "null")
        cell.font, cell.alignment, cell.border = DATA_FONT, CENTER_ALIGN, THIN_BORDER
        col += 1

        # 价格 ($)
        pv = product.get("price")
        pt = product.get("price_text")
        display_price = pt if pt else (f"${pv}" if pv else "null")
        cell = ws.cell(row=row_idx, column=col, value=display_price)
        cell.font, cell.alignment, cell.border = DATA_FONT, CENTER_ALIGN, THIN_BORDER
        col += 1

        # 评分
        v = product.get("rating")
        cell = ws.cell(row=row_idx, column=col, value=v if v else "null")
        cell.font, cell.alignment, cell.border = DATA_FONT, CENTER_ALIGN, THIN_BORDER
        col += 1

        # 评论数
        v = product.get("reviews")
        cell = ws.cell(row=row_idx, column=col, value=v if v else "null")
        cell.font, cell.alignment, cell.border = DATA_FONT, CENTER_ALIGN, THIN_BORDER
        col += 1

        # BSR排名
        v = product.get("bsr")
        cell = ws.cell(row=row_idx, column=col, value=v if v else "null")
        cell.font, cell.alignment, cell.border = DATA_FONT, LEFT_ALIGN, THIN_BORDER
        col += 1

        # 五点描述
        v = product.get("bullets")
        cell = ws.cell(row=row_idx, column=col, value=v if v else "null")
        cell.font, cell.alignment, cell.border = DATA_FONT, LEFT_ALIGN, THIN_BORDER
        col += 1

        # 规格参数
        v = product.get("specs")
        cell = ws.cell(row=row_idx, column=col, value=v if v else "null")
        cell.font, cell.alignment, cell.border = DATA_FONT, LEFT_ALIGN, THIN_BORDER
        col += 1

        # 月销量
        sv = product.get("monthly_sales")
        cell = ws.cell(row=row_idx, column=col, value=sv if sv else "null")
        cell.font, cell.alignment, cell.border = DATA_FONT, CENTER_ALIGN, THIN_BORDER
        col += 1

        # 客户差评（可选列）
        if with_reviews:
            v = product.get("negative_reviews")
            cell = ws.cell(row=row_idx, column=col, value=v if v else "null")
            cell.font, cell.alignment, cell.border = DATA_FONT, LEFT_ALIGN, THIN_BORDER
            col += 1

        # 商品链接
        url = product.get("url")
        if url:
            cell = ws.cell(row=row_idx, column=col, value=url)
            cell.hyperlink, cell.font = url, LINK_FONT
        else:
            cell = ws.cell(row=row_idx, column=col, value="null")
            cell.font = DATA_FONT
        cell.alignment, cell.border = LEFT_ALIGN, THIN_BORDER

    # ── 采集信息表 ───────────────────────────────────────────
    info_ws = wb.create_sheet("采集信息", 1)
    info_data = [
        ["采集时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["搜索关键词", keyword],
        ["采集数量", len(products)],
        ["版本", "v6.3 精准价格+评论采集版"],
        [],
        ["字段说明", ""],
        ["1. 首图", "高清商品首图（本地PNG）"],
        ["2. ASIN", "亚马逊标准识别码"],
        ["3. 商品标题", "完整商品名称"],
        ["4. 品牌名称", "Brand Name"],
        ["5. 价格($)", "BuyBox 主价格（TierP/0a/0b+精准提取，排除划线价）"],
        ["6. 评分", "星级评分"],
        ["7. 评论数", "累计评论数量"],
        ["8. BSR排名", "Best Sellers Rank 全量（最多8条）"],
        ["9. 五点描述", "About this item Bullet Points"],
        ["10. 规格参数", "仅含: Weight / Size / Material"],
        ["11. 月销量", "过去30天购买量"],
        ["12. 客户差评", "1-2星差评内容（前3条摘要，每条前150字）"] if with_reviews else ["12. 商品链接", "Amazon.com URL"],
        ["13. 商品链接", "Amazon.com URL"] if with_reviews else [],
        [],
        ["v6.3 更新", ""],
        ["价格修复", "新增 TierP 层(corePriceDisplay/priceToPay)；.a-price .a-offscreen 限制在 BuyBox 容器内；Tier3 改中位数策略"],
        ["评论采集", "新增差评采集（/product-reviews/ASIN/?filterByStar=critical）；每商品采2页=约20条；支持1星/2星/差评/全部/好评 5种筛选"],
    ]
    for row in info_data:
        if row:
            info_ws.append(row)
        else:
            info_ws.append([])

    wb.save(excel_path)
    return excel_path


# ════════════════════════════════════════════════════════════
#  主程序
# ════════════════════════════════════════════════════════════

def main():
    """
    用法：
      python amazon_scraper_core.py <搜索词> <排序> <数量> [输出目录] [是否采集差评] [差评页数]

    示例：
      # 仅采集商品数据（不含评论）
      python amazon_scraper_core.py "wireless earbuds" sales 20

      # 采集商品数据 + 差评（每商品2页=约20条差评）
      python amazon_scraper_core.py "wireless earbuds" sales 20 "" reviews 2

    参数说明：
      [4] 输出目录：可选，未指定或填 "" 时默认桌面日期文件夹
      [5] 评论模式：reviews（差评，默认） | all（全部） | skip（不采集）
      [6] 差评页数：每商品采几页，每页10条，默认2页
    """
    if len(sys.argv) < 4:
        print("用法: python amazon_scraper_core.py <搜索词> <排序方式> <数量> [输出目录] [评论模式] [差评页数]")
        print("排序方式: sales(销量) 或 reviews(评论数)")
        print("评论模式: reviews(差评,默认) | all(全部) | skip(不采集)")
        print("示例: python amazon_scraper_core.py \"yoga mat\" sales 20 \"\" reviews 2")
        return

    keyword = sys.argv[1]
    sort_by = sys.argv[2] if len(sys.argv) > 2 else "sales"
    total = int(sys.argv[3]) if len(sys.argv) > 3 else 10

    # 输出目录
    if len(sys.argv) > 4 and sys.argv[4].strip():
        output_dir = sys.argv[4].strip()
    else:
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        date_folder = datetime.now().strftime("%Y-%m-%d")
        output_dir = os.path.join(desktop, date_folder)

    # 评论采集参数
    review_mode = sys.argv[5].strip().lower() if len(sys.argv) > 5 else "skip"
    review_pages = int(sys.argv[6]) if len(sys.argv) > 6 else 2

    fetch_reviews_flag = review_mode not in ("skip", "no", "false", "0")
    review_filter = "negative" if review_mode in ("reviews", "negative", "critical") else (
        "all" if review_mode == "all" else "negative"
    )

    print(f"\n{'='*60}")
    print(f"  Amazon 商品采集器 v6.3 (精准价格+评论采集版)")
    print(f"{'='*60}")
    print(f"  搜索关键词: {keyword}")
    print(f"  排序方式: {sort_by}")
    print(f"  采集数量: {total}")
    print(f"  输出目录: {output_dir}")
    print(f"  评论采集: {'是（' + review_filter + '，每商品' + str(review_pages) + '页≈' + str(review_pages*10) + '条）' if fetch_reviews_flag else '否'}")
    print(f"{'='*60}\n")

    os.makedirs(output_dir, exist_ok=True)

    scraper = AmazonScraper()
    products = scraper.collect(
        keyword, sort_by, total,
        fetch_reviews_flag=fetch_reviews_flag,
        review_filter=review_filter,
        review_pages=review_pages
    )

    if not products:
        print("\n  [错误] 未能采集到任何商品数据")
        return

    print(f"\n  [完成] 成功采集 {len(products)} 个商品\n")
    print(f"  [导出] 正在生成Excel文件...")

    excel_path = export_excel(products, output_dir, keyword, with_reviews=fetch_reviews_flag)
    print(f"  [完成] Excel文件: {excel_path}")

    # 统计
    print(f"\n{'='*60}")
    print(f"  数据统计 (v6.3)")
    print(f"{'='*60}")

    fields_stats = [
        ("商品标题", "title"),       ("商品价格", "price"),
        ("评论数",   "reviews"),     ("月销量",   "monthly_sales"),
        ("ASIN",    "asin"),         ("评分",     "rating"),
        ("BSR排名",  "bsr"),
        ("五点描述","bullets"),      ("规格参数", "specs"),
        ("品牌名称","brand_name"),
    ]
    if fetch_reviews_flag:
        fields_stats.append(("客户差评", "negative_reviews"))

    for label, key in fields_stats:
        count = sum(1 for p in products if p.get(key))
        pct = count * 100 // len(products) if products else 0
        print(f"  {label}: {count}/{len(products)} ({pct}%)")
    print(f"{'='*60}\n")

    return excel_path


if __name__ == "__main__":
    main()
